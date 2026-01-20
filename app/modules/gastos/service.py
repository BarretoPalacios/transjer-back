from typing import List, Optional, Dict, Any
from bson import ObjectId
from app.modules.utils.core.code_generator.code_generator import generate_sequential_code
from app.core.database import get_database
from app.modules.gastos.model import Gasto
from app.modules.gastos.schema import GastoCreate, GastoUpdate, GastoFilter
from datetime import datetime
import pandas as pd
from io import BytesIO
import logging
import json
from math import ceil
import re

logger = logging.getLogger(__name__)

class GastoService:
    def __init__(self, db):
        self.db = db 
        self.collection = db["gastos"]
    
    def create_gasto(self, gasto_data: dict) -> dict:
        try:
            codigo_gasto = generate_sequential_code(
                counters_collection=self.db["counters"],
                target_collection=self.collection,
                sequence_name="gastos",
                field_name="id_gasto",
                prefix="GST-",
                length=10
            )

            gasto_data["id_gasto"] = codigo_gasto

            if not gasto_data.get("fecha_gasto"):
                gasto_data["fecha_gasto"] = datetime.now()

            gasto_model = Gasto(**gasto_data)

            result = self.collection.insert_one(
                gasto_model.model_dump(by_alias=True)
            )

            created_gasto = self.collection.find_one(
                {"_id": result.inserted_id}
            )
            created_gasto["id"] = str(created_gasto["_id"])
            del created_gasto["_id"]
            created_gasto["total"] = sum(d.get("valor", 0) for d in created_gasto.get("detalles_gastos", []))

            return created_gasto

        except Exception as e:
            logger.error(f"Error al crear gasto: {str(e)}")
            raise
    
    def get_gasto_by_id(self, gasto_id: str) -> Optional[dict]:
        try:
            if not ObjectId.is_valid(gasto_id):
                return None
            
            gasto = self.collection.find_one({"_id": ObjectId(gasto_id)})
            if gasto:
                gasto["id"] = str(gasto["_id"])
                del gasto["_id"]
                gasto["total"] = sum(d.get("valor", 0) for d in gasto.get("detalles_gastos", []))
            return gasto
            
        except Exception as e:
            logger.error(f"Error al obtener gasto: {str(e)}")
            return None
    
    def get_gasto_by_codigo(self, id_gasto: str) -> Optional[dict]:
        try:
            gasto = self.collection.find_one({"id_gasto": id_gasto})
            if gasto:
                gasto["id"] = str(gasto["_id"])
                del gasto["_id"]
                gasto["total"] = sum(d.get("valor", 0) for d in gasto.get("detalles_gastos", []))
            return gasto
            
        except Exception as e:
            logger.error(f"Error al obtener gasto por código: {str(e)}")
            return None
    
    def get_all_gastos(
            self,
            filter_params: Optional[GastoFilter] = None,
            page: int = 1,
            page_size: int = 10
        ) -> dict:
            try:
                query = {}

                if filter_params:
                    if filter_params.id_gasto:
                        query["id_gasto"] = safe_regex(filter_params.id_gasto)

                    if filter_params.placa:
                        query["placa"] = safe_regex(filter_params.placa)

                    if filter_params.ambito:
                        query["ambito"] = filter_params.ambito

                    if filter_params.estado:
                        query["estado"] = filter_params.estado

                    if filter_params.tipo_gasto:
                        query["detalles_gastos.tipo_gasto"] = filter_params.tipo_gasto

                    if filter_params.fecha_gasto_desde:
                        query["fecha_gasto"] = {"$gte": filter_params.fecha_gasto_desde}

                    if filter_params.fecha_gasto_hasta:
                        if "fecha_gasto" in query:
                            query["fecha_gasto"]["$lte"] = filter_params.fecha_gasto_hasta
                        else:
                            query["fecha_gasto"] = {"$lte": filter_params.fecha_gasto_hasta}

                    if filter_params.valor_minimo is not None:
                        query["detalles_gastos.valor"] = {"$gte": filter_params.valor_minimo}

                    if filter_params.valor_maximo is not None:
                        if "detalles_gastos.valor" in query:
                            query["detalles_gastos.valor"]["$lte"] = filter_params.valor_maximo
                        else:
                            query["detalles_gastos.valor"] = {"$lte": filter_params.valor_maximo}

                query = {k: v for k, v in query.items() if v is not None}

                total = self.collection.count_documents(query)

                skip = (page - 1) * page_size

                gastos = list(
                    self.collection.find(query)
                    .sort("fecha_gasto", -1)
                    .skip(skip)
                    .limit(page_size)
                )

                for gasto in gastos:
                    gasto["id"] = str(gasto["_id"])
                    del gasto["_id"]
                    gasto["total"] = sum(d.get("valor", 0) for d in gasto.get("detalles_gastos", []))

                total_pages = ceil(total / page_size) if page_size > 0 else 0

                return {
                    "items": gastos,
                    "total": total,
                    "page": page,
                    "page_size": page_size,
                    "total_pages": total_pages,
                    "has_next": page < total_pages,
                    "has_prev": page > 1
                }

            except Exception as e:
                logger.error(f"Error al obtener gastos: {str(e)}")
                raise
    
    def update_gasto(self, gasto_id: str, update_data: dict) -> Optional[dict]:
        try:
            if not ObjectId.is_valid(gasto_id):
                return None
            
            update_dict = {k: v for k, v in update_data.items() if v is not None}
            
            if not update_dict:
                return self.get_gasto_by_id(gasto_id)
            
            self.collection.update_one(
                {"_id": ObjectId(gasto_id)},
                {"$set": update_dict}
            )
            
            return self.get_gasto_by_id(gasto_id)
            
        except Exception as e:
            logger.error(f"Error al actualizar gasto: {str(e)}")
            raise
    
    def delete_gasto(self, gasto_id: str) -> bool:
        try:
            if not ObjectId.is_valid(gasto_id):
                return False
            
            result = self.collection.delete_one({"_id": ObjectId(gasto_id)})
            return result.deleted_count > 0
            
        except Exception as e:
            logger.error(f"Error al eliminar gasto: {str(e)}")
            return False
    
    def export_to_excel(self, filter_params: Optional[GastoFilter] = None) -> BytesIO:
        try:
            gastos = self.get_all_gastos_sin_paginacion(filter_params)
            
            if not gastos:
                df = pd.DataFrame(columns=[
                    "ID", "ID Gasto", "Placa", "Ámbito", "Fecha Gasto",
                    "Estado", "Total", "Usuario Registro", "Fecha Registro",
                    "Detalles Gastos"
                ])
            else:
                excel_data = []
                for gasto in gastos:
                    detalles_str = ""
                    if gasto.get("detalles_gastos"):
                        detalles_str = json.dumps(gasto["detalles_gastos"], ensure_ascii=False)
                    
                    total = sum(d.get("valor", 0) for d in gasto.get("detalles_gastos", []))
                    
                    excel_data.append({
                        "ID": gasto.get("id", ""),
                        "ID Gasto": gasto.get("id_gasto", ""),
                        "Placa": gasto.get("placa", ""),
                        "Ámbito": gasto.get("ambito", ""),
                        "Fecha Gasto": gasto.get("fecha_gasto", "").strftime("%Y-%m-%d %H:%M:%S") if gasto.get("fecha_gasto") else "",
                        "Estado": gasto.get("estado", ""),
                        "Total": total,
                        "Usuario Registro": gasto.get("usuario_registro", ""),
                        "Fecha Registro": gasto.get("fecha_registro", "").strftime("%Y-%m-%d %H:%M:%S") if gasto.get("fecha_registro") else "",
                        "Detalles Gastos": detalles_str
                    })
                
                df = pd.DataFrame(excel_data)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Gastos')
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {str(e)}")
            raise
    
    def import_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        try:
            import io
            
            df = pd.read_excel(io.BytesIO(file_content))
            
            created = 0
            updated = 0
            errors = []
            skipped = 0
            
            for index, row in df.iterrows():
                try:
                    gasto_data = {
                        "placa": str(row.get("Placa", "")).strip(),
                        "ambito": str(row.get("Ámbito", "local")).strip().lower(),
                        "estado": str(row.get("Estado", "pendiente")).strip().lower(),
                        "fecha_registro": datetime.now()
                    }
                    
                    fecha_gasto_str = str(row.get("Fecha Gasto", "")).strip()
                    if fecha_gasto_str and fecha_gasto_str not in ["", "nan", "None"]:
                        try:
                            gasto_data["fecha_gasto"] = pd.to_datetime(fecha_gasto_str)
                        except:
                            gasto_data["fecha_gasto"] = datetime.now()
                    else:
                        gasto_data["fecha_gasto"] = datetime.now()
                    
                    detalles_str = str(row.get("Detalles Gastos", "")).strip()
                    if detalles_str and detalles_str not in ["", "nan", "None"]:
                        try:
                            gasto_data["detalles_gastos"] = json.loads(detalles_str)
                        except:
                            errors.append(f"Fila {index + 2}: Error al parsear detalles de gastos")
                            continue
                    else:
                        errors.append(f"Fila {index + 2}: Detalles de gastos es requerido")
                        continue
                    
                    if not gasto_data.get("placa"):
                        errors.append(f"Fila {index + 2}: Placa es requerida")
                        continue
                    
                    if not gasto_data.get("detalles_gastos"):
                        errors.append(f"Fila {index + 2}: Debe incluir al menos un detalle de gasto")
                        continue
                    
                    if gasto_data.get("estado") not in ["pendiente", "aprobado", "rechazado", "pagado"]:
                        gasto_data["estado"] = "pendiente"
                    
                    if gasto_data.get("ambito") not in ["local", "nacional"]:
                        gasto_data["ambito"] = "local"
                    
                    id_gasto_excel = str(row.get("ID Gasto", "")).strip()
                    if id_gasto_excel and id_gasto_excel not in ["", "nan", "None"]:
                        existing = self.collection.find_one({"id_gasto": id_gasto_excel})
                        if existing:
                            gasto_data.pop("fecha_registro", None)
                            self.collection.update_one(
                                {"id_gasto": id_gasto_excel},
                                {"$set": gasto_data}
                            )
                            updated += 1
                            continue
                    
                    self.create_gasto(gasto_data)
                    created += 1
                        
                except Exception as e:
                    errors.append(f"Fila {index + 2}: {str(e)}")
                    continue
            
            return {
                "total_rows": len(df),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "has_errors": len(errors) > 0,
                "success_rate": f"{((created + updated) / len(df) * 100):.1f}%" if len(df) > 0 else "0%"
            }
            
        except Exception as e:
            logger.error(f"Error al importar desde Excel: {str(e)}")
            raise

    def get_stats(self) -> Dict[str, Any]:
        try:
            total = self.collection.count_documents({})
            pendientes = self.collection.count_documents({"estado": "pendiente"})
            aprobados = self.collection.count_documents({"estado": "aprobado"})
            rechazados = self.collection.count_documents({"estado": "rechazado"})
            pagados = self.collection.count_documents({"estado": "pagado"})
            
            pipeline_ambito = [
                {"$group": {"_id": "$ambito", "count": {"$sum": 1}}}
            ]
            
            ambitos = {}
            for result in self.collection.aggregate(pipeline_ambito):
                ambitos[result["_id"]] = result["count"]
            
            pipeline_placa = [
                {"$group": {"_id": "$placa", "count": {"$sum": 1}}}
            ]
            
            placas = {}
            for result in self.collection.aggregate(pipeline_placa):
                placas[result["_id"]] = result["count"]
            
            pipeline_tipo = [
                {"$unwind": "$detalles_gastos"},
                {"$group": {"_id": "$detalles_gastos.tipo_gasto", "count": {"$sum": 1}}}
            ]
            
            tipos = {}
            for result in self.collection.aggregate(pipeline_tipo):
                tipos[result["_id"]] = result["count"]
            
            pipeline_total = [
                {"$unwind": "$detalles_gastos"},
                {"$group": {"_id": None, "total": {"$sum": "$detalles_gastos.valor"}}}
            ]
            
            total_gastado = 0
            for result in self.collection.aggregate(pipeline_total):
                total_gastado = result.get("total", 0)
            
            return {
                "total": total,
                "pendientes": pendientes,
                "aprobados": aprobados,
                "rechazados": rechazados,
                "pagados": pagados,
                "por_ambito": ambitos,
                "por_placa": placas,
                "por_tipo_gasto": tipos,
                "total_gastado": total_gastado
            }
            
        except Exception as e:
            logger.error(f"Error al obtener estadísticas: {str(e)}")
            return {}

    def get_all_gastos_sin_paginacion(
            self,
            filter_params: Optional[GastoFilter] = None
        ) -> List[dict]:
            try:
                query = {}

                if filter_params:
                    if filter_params.id_gasto:
                        query["id_gasto"] = safe_regex(filter_params.id_gasto)

                    if filter_params.placa:
                        query["placa"] = safe_regex(filter_params.placa)

                    if filter_params.ambito:
                        query["ambito"] = filter_params.ambito

                    if filter_params.estado:
                        query["estado"] = filter_params.estado

                    if filter_params.tipo_gasto:
                        query["detalles_gastos.tipo_gasto"] = filter_params.tipo_gasto

                    if filter_params.fecha_gasto_desde:
                        query["fecha_gasto"] = {"$gte": filter_params.fecha_gasto_desde}

                    if filter_params.fecha_gasto_hasta:
                        if "fecha_gasto" in query:
                            query["fecha_gasto"]["$lte"] = filter_params.fecha_gasto_hasta
                        else:
                            query["fecha_gasto"] = {"$lte": filter_params.fecha_gasto_hasta}

                query = {k: v for k, v in query.items() if v is not None}

                gastos = list(
                    self.collection
                    .find(query)
                    .sort("fecha_gasto", -1)
                )

                for gasto in gastos:
                    gasto["id"] = str(gasto["_id"])
                    del gasto["_id"]

                return gastos

            except Exception as e:
                logger.error(f"Error al obtener gastos sin paginación: {str(e)}")
                raise

    def generate_excel_template(self) -> BytesIO:
        try:
            template_data = [{
                "Placa": "ABC-123",
                "Ámbito": "local",
                "Fecha Gasto": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Estado": "pendiente",
                "Detalles Gastos": json.dumps([
                    {
                        "tipo_gasto": "Combustible",
                        "tipo_gasto_personalizado": None,
                        "valor": 150.50,
                        "observacion": "Tanque lleno"
                    }
                ], ensure_ascii=False)
            }]
            
            df = pd.DataFrame(template_data)
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Gastos')
                
                workbook = writer.book
                worksheet = writer.sheets['Gastos']
                
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                column_widths = {
                    'A': 15,
                    'B': 15,
                    'C': 25,
                    'D': 15,
                    'E': 60
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                instructions_data = {
                    "Campo": [
                        "Placa",
                        "Ámbito",
                        "Fecha Gasto",
                        "Estado",
                        "Detalles Gastos"
                    ],
                    "Obligatorio": [
                        "SÍ", "SÍ", "NO", "NO", "SÍ"
                    ],
                    "Descripción": [
                        "Placa del vehículo (6-7 caracteres)",
                        "Ámbito del gasto: local, nacional",
                        "Fecha del gasto (formato: YYYY-MM-DD HH:MM:SS)",
                        "Estado: pendiente, aprobado, rechazado, pagado",
                        "JSON con array de detalles: tipo_gasto, valor, observacion"
                    ],
                    "Ejemplo": [
                        "ABC-123",
                        "local",
                        "2025-01-19 08:30:00",
                        "pendiente",
                        '[{"tipo_gasto":"Combustible","valor":150.50,"observacion":"Tanque lleno"}]'
                    ]
                }
                
                df_instructions = pd.DataFrame(instructions_data)
                df_instructions.to_excel(writer, sheet_name='Instrucciones', index=False)
                
                ws_instructions = writer.sheets['Instrucciones']
                
                for cell in ws_instructions[1]:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                ws_instructions.column_dimensions['A'].width = 25
                ws_instructions.column_dimensions['B'].width = 15
                ws_instructions.column_dimensions['C'].width = 50
                ws_instructions.column_dimensions['D'].width = 60
                
                for row in ws_instructions.iter_rows(min_row=2, max_row=ws_instructions.max_row):
                    ws_instructions.row_dimensions[row[0].row].height = 30
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al generar plantilla Excel: {str(e)}")
            raise

    def get_gastos_by_placa(self, placa: str) -> List[dict]:
        try:
            gastos = list(
                self.collection
                .find({"placa": placa})
                .sort("fecha_gasto", -1)
            )

            for gasto in gastos:
                gasto["id"] = str(gasto["_id"])
                del gasto["_id"]
                gasto["total"] = sum(d.get("valor", 0) for d in gasto.get("detalles_gastos", []))

            return gastos
        except Exception as e:
            logger.error(f"Error al obtener gastos por placa: {str(e)}")
            return []


def safe_regex(value: str):
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    return {"$regex": re.escape(value), "$options": "i"}