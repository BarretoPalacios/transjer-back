from typing import List, Optional, Dict, Any
from bson import ObjectId
from app.modules.utils.core.code_generator.code_generator import generate_sequential_code
from app.core.database import get_database
from app.modules.dataservice.models.flota import Flota
from app.modules.dataservice.schemas.flota_schema import FlotaCreate, FlotaUpdate, FlotaFilter
from datetime import datetime, date
import pandas as pd
from io import BytesIO
import logging
import json
from math import ceil
import re

logger = logging.getLogger(__name__)

class FlotaService:
    def __init__(self, db):
        self.db = db 
        self.collection = db["flota"]
    
    def _convert_dates_to_datetime(self, data: dict) -> dict:
        """Convertir campos date a datetime para MongoDB"""
        date_fields = [
            'revision_tecnica_emision',
            'revision_tecnica_vencimiento', 
            'soat_vigencia_inicio',
            'soat_vigencia_fin',
            'extintor_vencimiento'
        ]
        
        for field in date_fields:
            if field in data and data[field] is not None:
                if isinstance(data[field], date) and not isinstance(data[field], datetime):
                    # Convertir date a datetime (a medianoche)
                    data[field] = datetime.combine(data[field], datetime.min.time())
        
        return data
    
    def _convert_dates_to_date(self, flota: dict) -> dict:
        """Convertir campos datetime a date para respuesta"""
        date_fields = [
            'revision_tecnica_emision',
            'revision_tecnica_vencimiento', 
            'soat_vigencia_inicio',
            'soat_vigencia_fin',
            'extintor_vencimiento'
        ]
        
        for field in date_fields:
            if field in flota and flota[field] is not None:
                if isinstance(flota[field], datetime):
                    flota[field] = flota[field].date()
        
        return flota
    
    def _prepare_flota_response(self, flota: dict) -> dict:
        """Preparar respuesta de flota convirtiendo fechas y ObjectId"""
        if flota:
            flota["id"] = str(flota["_id"])
            del flota["_id"]
            flota = self._convert_dates_to_date(flota)
        return flota
    
    def create_flota(self, flota_data: dict) -> dict:
        try:
            # 1️⃣ Verificar placa (regla de negocio)
            existing_placa = self.collection.find_one({
                "placa": flota_data["placa"]
            })
            if existing_placa:
                raise ValueError(
                    f"La placa {flota_data['placa']} ya está registrada"
                )

            # 2️⃣ Si viene número de licencia, verificar que no exista duplicado
            if flota_data.get("numero_licencia"):
                existing_licencia = self.collection.find_one({
                    "numero_licencia": flota_data["numero_licencia"]
                })
                if existing_licencia:
                    raise ValueError(
                        f"El número de licencia {flota_data['numero_licencia']} ya está registrado"
                    )

            # 3️⃣ Generar código automáticamente
            codigo_flota = generate_sequential_code(
                counters_collection=self.db["counters"],
                target_collection=self.collection,
                sequence_name="flota",
                field_name="codigo_flota",
                prefix="FL-", 
                length=10
            )

            flota_data["codigo_flota"] = codigo_flota

            # 4️⃣ Convertir dates a datetime antes de guardar
            flota_data = self._convert_dates_to_datetime(flota_data)

            # 5️⃣ Crear modelo
            flota_model = Flota(**flota_data)

            # 6️⃣ Insertar
            result = self.collection.insert_one(
                flota_model.model_dump(by_alias=True)
            )

            # 7️⃣ Retornar creado
            created_flota = self.collection.find_one(
                {"_id": result.inserted_id}
            )
            return self._prepare_flota_response(created_flota)

        except Exception as e:
            logger.error(f"Error al crear vehículo: {str(e)}")
            raise
    
    def get_flota_by_id(self, flota_id: str) -> Optional[dict]:
        """Obtener vehículo por ID"""
        try:
            if not ObjectId.is_valid(flota_id):
                return None
            
            flota = self.collection.find_one({"_id": ObjectId(flota_id)})
            if flota:
                return self._prepare_flota_response(flota)
            return None
            
        except Exception as e:
            logger.error(f"Error al obtener vehículo: {str(e)}")
            return None
    
    def get_flota_by_codigo(self, codigo_flota: str) -> Optional[dict]:
        """Obtener vehículo por código"""
        try:
            flota = self.collection.find_one({"codigo_flota": codigo_flota})
            if flota:
                return self._prepare_flota_response(flota)
            return None
            
        except Exception as e:
            logger.error(f"Error al obtener vehículo por código: {str(e)}")
            return None
    
    def get_flota_by_placa(self, placa: str) -> Optional[dict]:
        """Obtener vehículo por placa"""
        try:
            flota = self.collection.find_one({"placa": placa})
            if flota:
                return self._prepare_flota_response(flota)
            return None
            
        except Exception as e:
            logger.error(f"Error al obtener vehículo por placa: {str(e)}")
            return None
    
    def get_flota_by_licencia(self, numero_licencia: str) -> Optional[dict]:
        """Obtener vehículo por número de licencia"""
        try:
            flota = self.collection.find_one({"numero_licencia": numero_licencia})
            if flota:
                return self._prepare_flota_response(flota)
            return None
            
        except Exception as e:
            logger.error(f"Error al obtener vehículo por licencia: {str(e)}")
            return None
    
    def get_all_flotas(
            self,
            filter_params: Optional[FlotaFilter] = None,
            page: int = 1,
            page_size: int = 10
        ) -> dict:
            """Obtener todos los vehículos con filtros opcionales y paginación"""
            try:
                query = {}

                if filter_params:
                    if filter_params.codigo_flota:
                        query["codigo_flota"] = safe_regex(filter_params.codigo_flota)

                    if filter_params.placa:
                        query["placa"] = safe_regex(filter_params.placa)

                    if filter_params.marca:
                        query["marca"] = safe_regex(filter_params.marca)

                    if filter_params.modelo:
                        query["modelo"] = safe_regex(filter_params.modelo)

                    if filter_params.anio:
                        query["anio"] = filter_params.anio

                    if filter_params.tipo_vehiculo:
                        query["tipo_vehiculo"] = filter_params.tipo_vehiculo

                    if filter_params.tipo_combustible:
                        query["tipo_combustible"] = filter_params.tipo_combustible

                    # Nuevos filtros para campos de conductor
                    if filter_params.nombre_conductor:
                        query["nombre_conductor"] = safe_regex(filter_params.nombre_conductor)

                    if filter_params.numero_licencia:
                        query["numero_licencia"] = safe_regex(filter_params.numero_licencia)

                    if filter_params.mtc_numero:
                        query["mtc_numero"] = safe_regex(filter_params.mtc_numero)

                    if filter_params.activo is not None:
                        query["activo"] = filter_params.activo

                # Limpiar filtros nulos
                query = {k: v for k, v in query.items() if v is not None}

                total = self.collection.count_documents(query)

                skip = (page - 1) * page_size

                flotas = list(
                    self.collection.find(query)
                    .sort("placa", 1)
                    .skip(skip)
                    .limit(page_size)
                )

                # Preparar cada flota individualmente
                flotas_preparadas = []
                for flota in flotas:
                    flota_preparada = self._prepare_flota_response(flota)
                    flotas_preparadas.append(flota_preparada)

                total_pages = ceil(total / page_size) if page_size > 0 else 0

                return {
                    "items": flotas_preparadas,
                    "total": total,
                    "page": page,
                    "page_size": page_size,
                    "total_pages": total_pages,
                    "has_next": page < total_pages,
                    "has_prev": page > 1
                }

            except Exception as e:
                logger.error(f"Error al obtener vehículos: {str(e)}")
                raise
    
    def update_flota(self, flota_id: str, update_data: dict) -> Optional[dict]:
        """Actualizar vehículo"""
        try:
            if not ObjectId.is_valid(flota_id):
                return None
            
            # Filtrar campos None
            update_dict = {k: v for k, v in update_data.items() if v is not None}
            
            if not update_dict:
                return self.get_flota_by_id(flota_id)
            
            # Si se actualiza la placa, verificar que no exista otra con la misma
            if "placa" in update_dict:
                existing = self.collection.find_one({
                    "placa": update_dict["placa"],
                    "_id": {"$ne": ObjectId(flota_id)}
                }) 
                
                if existing:
                    raise ValueError(f"Ya existe un vehículo con la placa {update_dict['placa']}")
            
            # Si se actualiza número de licencia, verificar que no exista duplicado
            if "numero_licencia" in update_dict and update_dict["numero_licencia"]:
                existing_licencia = self.collection.find_one({
                    "numero_licencia": update_dict["numero_licencia"],
                    "_id": {"$ne": ObjectId(flota_id)}
                })
                
                if existing_licencia:
                    raise ValueError(f"El número de licencia {update_dict['numero_licencia']} ya está registrado en otro vehículo")
            
            # Convertir dates a datetime antes de guardar
            update_dict = self._convert_dates_to_datetime(update_dict)
            
            # Actualizar en base de datos
            self.collection.update_one(
                {"_id": ObjectId(flota_id)},
                {"$set": update_dict}
            )
            
            return self.get_flota_by_id(flota_id)
            
        except Exception as e:
            logger.error(f"Error al actualizar vehículo: {str(e)}")
            raise
    
    def delete_flota(self, flota_id: str) -> bool:
        """Eliminar vehículo"""
        try:
            if not ObjectId.is_valid(flota_id):
                return False
            
            result = self.collection.delete_one({"_id": ObjectId(flota_id)})
            return result.deleted_count > 0
            
        except Exception as e:
            logger.error(f"Error al eliminar vehículo: {str(e)}")
            return False
    
    def export_to_excel(self, filter_params: Optional[FlotaFilter] = None) -> BytesIO:
        """Exportar vehículos a Excel"""
        try:
            flotas = self.get_all_flotas_sin_paginacion(filter_params)
            
            if not flotas:
                # Crear DataFrame vacío
                df = pd.DataFrame(columns=[
                    "ID", "Código Flota", "Placa", "Marca", "Modelo", "Año",
                    "Toneladas (TN)", "Metraje (M3)", "Tipo Vehículo", 
                    "Tipo Combustible", "Nombre Conductor", "Número Licencia",
                    "Revisión Técnica Emisión", "Revisión Técnica Vencimiento", 
                    "SOAT Inicio", "SOAT Fin", "MTC Número", "Extintor Vencimiento",
                    "Cantidad Parihuelas", "Días Alerta RT", "Días Alerta SOAT",
                    "Días Alerta Extintor", "Observaciones", "Activo", "Fecha Registro"
                ])
            else:
                # Preparar datos para Excel
                excel_data = []
                for flota in flotas:
                    excel_data.append({
                        "ID": flota.get("id", ""),
                        "Código Flota": flota.get("codigo_flota", ""),
                        "Placa": flota.get("placa", ""),
                        "Marca": flota.get("marca", ""),
                        "Modelo": flota.get("modelo", ""),
                        "Año": flota.get("anio", ""),
                        "Toneladas (TN)": flota.get("tn", ""),
                        "Metraje (M3)": flota.get("m3", ""),
                        "Tipo Vehículo": flota.get("tipo_vehiculo", ""),
                        "Tipo Combustible": flota.get("tipo_combustible", ""),
                        "Nombre Conductor": flota.get("nombre_conductor", ""),
                        "Número Licencia": flota.get("numero_licencia", ""),
                        "Revisión Técnica Emisión": flota.get("revision_tecnica_emision", ""),
                        "Revisión Técnica Vencimiento": flota.get("revision_tecnica_vencimiento", ""),
                        "SOAT Inicio": flota.get("soat_vigencia_inicio", ""),
                        "SOAT Fin": flota.get("soat_vigencia_fin", ""),
                        "MTC Número": flota.get("mtc_numero", ""),
                        "Extintor Vencimiento": flota.get("extintor_vencimiento", ""),
                        "Cantidad Parihuelas": flota.get("cantidad_parihuelas", 0),
                        "Días Alerta RT": flota.get("dias_alerta_revision_tecnica", 30),
                        "Días Alerta SOAT": flota.get("dias_alerta_soat", 30),
                        "Días Alerta Extintor": flota.get("dias_alerta_extintor", 15),
                        "Observaciones": flota.get("observaciones", ""),
                        "Activo": "Sí" if flota.get("activo") else "No",
                        "Fecha Registro": flota.get("fecha_registro", "").strftime("%Y-%m-%d %H:%M:%S") if flota.get("fecha_registro") else ""
                    })
                
                df = pd.DataFrame(excel_data)
            
            # Crear Excel en memoria
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Flota')
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {str(e)}")
            raise
    
    def import_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Importar vehículos desde Excel (formato simplificado)"""
        try:
            import io
            
            df = pd.read_excel(io.BytesIO(file_content))
            
            created = 0
            updated = 0
            errors = []
            skipped = 0
            
            for index, row in df.iterrows():
                try:
                    # Función auxiliar para parsear fechas
                    def parse_date(value):
                        if pd.isna(value) or value == "" or value is None:
                            return None
                        if isinstance(value, date):
                            return value
                        if isinstance(value, datetime):
                            return value.date()
                        try:
                            return pd.to_datetime(value).date()
                        except:
                            return None
                    
                    # Construir datos del vehículo
                    flota_data = {
                        "placa": str(row.get("Placa", "")).strip().upper(),
                        "marca": str(row.get("Marca", "")).strip(),
                        "modelo": str(row.get("Modelo", "")).strip(),
                        "anio": int(row.get("Año", 0)),
                        "tn": float(row.get("Toneladas (TN)", 0)),
                        "m3": float(row.get("Metraje (M3)", 0)),
                        "tipo_vehiculo": str(row.get("Tipo Vehículo", "")).strip(),
                        "fecha_registro": datetime.now()
                    }
                    
                    # Campos opcionales incluyendo los nuevos
                    campos_opcionales = {
                        "tipo_combustible": "Tipo Combustible",
                        "nombre_conductor": "Nombre Conductor",
                        "numero_licencia": "Número Licencia",
                        "mtc_numero": "MTC Número",
                        "cantidad_parihuelas": "Cantidad Parihuelas",
                        "dias_alerta_revision_tecnica": "Días Alerta RT",
                        "dias_alerta_soat": "Días Alerta SOAT",
                        "dias_alerta_extintor": "Días Alerta Extintor",
                        "observaciones": "Observaciones"
                    }
                    
                    for campo_db, campo_excel in campos_opcionales.items():
                        valor = row.get(campo_excel)
                        if pd.notna(valor) and str(valor).strip() and str(valor).strip().lower() != "nan":
                            valor_limpio = str(valor).strip()
                            
                            # Convertir campos numéricos
                            if campo_db in ["cantidad_parihuelas", "dias_alerta_revision_tecnica", "dias_alerta_soat", "dias_alerta_extintor"]:
                                try:
                                    flota_data[campo_db] = int(float(valor_limpio))
                                except (ValueError, TypeError):
                                    if campo_db == "cantidad_parihuelas":
                                        flota_data[campo_db] = 0
                                    elif campo_db == "dias_alerta_revision_tecnica" or campo_db == "dias_alerta_soat":
                                        flota_data[campo_db] = 30
                                    elif campo_db == "dias_alerta_extintor":
                                        flota_data[campo_db] = 15
                            else:
                                flota_data[campo_db] = valor_limpio
                    
                    # Campos de fecha
                    flota_data["revision_tecnica_emision"] = parse_date(row.get("Revisión Técnica Emisión"))
                    flota_data["revision_tecnica_vencimiento"] = parse_date(row.get("Revisión Técnica Vencimiento"))
                    flota_data["soat_vigencia_inicio"] = parse_date(row.get("SOAT Inicio"))
                    flota_data["soat_vigencia_fin"] = parse_date(row.get("SOAT Fin"))
                    flota_data["extintor_vencimiento"] = parse_date(row.get("Extintor Vencimiento"))
                    
                    # Campo activo
                    activo_val = str(row.get("Activo", "Sí")).strip().lower()
                    flota_data["activo"] = activo_val in ["sí", "si", "yes", "true", "1"]
                    
                    # Validar campos obligatorios
                    if not flota_data.get("placa"):
                        errors.append(f"Fila {index + 2}: Placa es requerida")
                        continue
                    
                    if not flota_data.get("marca"):
                        errors.append(f"Fila {index + 2}: Marca es requerida")
                        continue
                    
                    if not flota_data.get("modelo"):
                        errors.append(f"Fila {index + 2}: Modelo es requerido")
                        continue
                    
                    if not flota_data.get("anio") or flota_data["anio"] < 1990 or flota_data["anio"] > 2025:
                        errors.append(f"Fila {index + 2}: Año inválido (debe estar entre 1990 y 2025)")
                        continue
                    
                    if flota_data["tn"] < 0:
                        errors.append(f"Fila {index + 2}: Toneladas no puede ser negativo")
                        continue
                    
                    if flota_data["m3"] < 0:
                        errors.append(f"Fila {index + 2}: Metraje no puede ser negativo")
                        continue
                    
                    if not flota_data.get("tipo_vehiculo"):
                        errors.append(f"Fila {index + 2}: Tipo de vehículo es requerido")
                        continue
                    
                    # Verificar si ya existe por código (si viene en el Excel)
                    codigo_excel = str(row.get("Código Flota", "")).strip()
                    if codigo_excel and codigo_excel not in ["", "nan", "None"]:
                        existing = self.collection.find_one({"codigo_flota": codigo_excel})
                        if existing:
                            # Actualizar vehículo existente
                            flota_data.pop("fecha_registro", None)  # No actualizar fecha de registro
                            flota_data = self._convert_dates_to_datetime(flota_data)
                            self.collection.update_one(
                                {"codigo_flota": codigo_excel},
                                {"$set": flota_data}
                            )
                            updated += 1
                            continue
                    
                    # Verificar si existe por placa
                    existing_placa = self.collection.find_one({
                        "placa": flota_data["placa"]
                    })
                    
                    if existing_placa:
                        # Vehículo duplicado, lo saltamos
                        skipped += 1
                        errors.append(f"Fila {index + 2}: Vehículo duplicado - Placa {flota_data['placa']} ya existe")
                        continue
                    
                    # Verificar si existe por número de licencia (si se proporciona)
                    if flota_data.get("numero_licencia"):
                        existing_licencia = self.collection.find_one({
                            "numero_licencia": flota_data["numero_licencia"]
                        })
                        
                        if existing_licencia:
                            # Licencia duplicada, lo saltamos
                            skipped += 1
                            errors.append(f"Fila {index + 2}: Licencia duplicada - Número {flota_data['numero_licencia']} ya existe")
                            continue
                    
                    # Crear nuevo vehículo (sin código, se generará automáticamente)
                    self.create_flota(flota_data)
                    created += 1
                        
                except Exception as e:
                    errors.append(f"Fila {index + 2}: {str(e)}")
                    continue
            
            return {
                "total_rows": len(df),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "has_errors": len(errors) > 0,
                "success_rate": f"{((created + updated) / len(df) * 100):.1f}%" if len(df) > 0 else "0%"
            }
            
        except Exception as e:
            logger.error(f"Error al importar desde Excel: {str(e)}")
            raise

    def get_stats(self) -> Dict[str, Any]:
        """Obtener estadísticas de flota"""
        try:
            total = self.collection.count_documents({})
            activos = self.collection.count_documents({"activo": True})
            inactivos = self.collection.count_documents({"activo": False})
            
            # Agrupar por tipo de vehículo
            pipeline_tipo = [
                {"$group": {"_id": "$tipo_vehiculo", "count": {"$sum": 1}}}
            ]
            
            tipos = {}
            for result in self.collection.aggregate(pipeline_tipo):
                tipos[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Agrupar por marca
            pipeline_marca = [
                {"$group": {"_id": "$marca", "count": {"$sum": 1}}}
            ]
            
            marcas = {}
            for result in self.collection.aggregate(pipeline_marca):
                marcas[result["_id"]] = result["count"]
            
            # Agrupar por tipo de combustible
            pipeline_combustible = [
                {"$group": {"_id": "$tipo_combustible", "count": {"$sum": 1}}}
            ]
            
            combustibles = {}
            for result in self.collection.aggregate(pipeline_combustible):
                combustibles[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Estadísticas de conductores
            conductores_con_info = self.collection.count_documents({
                "nombre_conductor": {"$ne": None, "$exists": True}
            })
            
            licencias_registradas = self.collection.count_documents({
                "numero_licencia": {"$ne": None, "$exists": True}
            })
            
            # Capacidad total
            pipeline_capacidad = [
                {"$group": {
                    "_id": None,
                    "total_tn": {"$sum": "$tn"},
                    "total_m3": {"$sum": "$m3"},
                    "total_parihuelas": {"$sum": "$cantidad_parihuelas"}
                }}
            ]
            
            capacidad = {"total_tn": 0, "total_m3": 0, "total_parihuelas": 0}
            for result in self.collection.aggregate(pipeline_capacidad):
                capacidad["total_tn"] = result.get("total_tn", 0)
                capacidad["total_m3"] = result.get("total_m3", 0)
                capacidad["total_parihuelas"] = result.get("total_parihuelas", 0)
            
            # Documentos por vencer (próximos 30 días)
            documentos_vencidos = self.get_flotas_con_documentos_vencidos(30)
            
            return {
                "total": total,
                "activos": activos,
                "inactivos": inactivos,
                "conductores_con_info": conductores_con_info,
                "licencias_registradas": licencias_registradas,
                "por_tipo_vehiculo": tipos,
                "por_marca": marcas,
                "por_tipo_combustible": combustibles,
                "capacidad_total": capacidad,
                "alertas_vencimiento": {
                    "revision_tecnica": len(documentos_vencidos["revision_tecnica"]),
                    "soat": len(documentos_vencidos["soat"]),
                    "extintor": len(documentos_vencidos["extintor"])
                }
            }
            
        except Exception as e:
            logger.error(f"Error al obtener estadísticas: {str(e)}")
            return {}

    def get_all_flotas_sin_paginacion(
            self,
            filter_params: Optional[FlotaFilter] = None
        ) -> List[dict]:
            """Obtener TODOS los vehículos sin paginación (para exportación)"""
            try:
                query = {}

                if filter_params:
                    if filter_params.codigo_flota:
                        query["codigo_flota"] = safe_regex(filter_params.codigo_flota)

                    if filter_params.placa:
                        query["placa"] = safe_regex(filter_params.placa)

                    if filter_params.marca:
                        query["marca"] = safe_regex(filter_params.marca)

                    if filter_params.modelo:
                        query["modelo"] = safe_regex(filter_params.modelo)

                    if filter_params.anio:
                        query["anio"] = filter_params.anio

                    if filter_params.tipo_vehiculo:
                        query["tipo_vehiculo"] = filter_params.tipo_vehiculo

                    if filter_params.tipo_combustible:
                        query["tipo_combustible"] = filter_params.tipo_combustible

                    # Nuevos filtros para campos de conductor
                    if filter_params.nombre_conductor:
                        query["nombre_conductor"] = safe_regex(filter_params.nombre_conductor)

                    if filter_params.numero_licencia:
                        query["numero_licencia"] = safe_regex(filter_params.numero_licencia)

                    if filter_params.mtc_numero:
                        query["mtc_numero"] = safe_regex(filter_params.mtc_numero)

                    if filter_params.activo is not None:
                        query["activo"] = filter_params.activo

                # Limpiar filtros None
                query = {k: v for k, v in query.items() if v is not None}

                flotas = list(
                    self.collection
                    .find(query)
                    .sort("placa", 1)
                )

                return [self._prepare_flota_response(flota) for flota in flotas]

            except Exception as e:
                logger.error(f"Error al obtener vehículos (sin paginación): {str(e)}")
                raise

    def generate_excel_template(self) -> BytesIO:
        """Generar plantilla de Excel vacía para importación"""
        try:
            # Crear DataFrame con columnas y una fila de ejemplo
            template_data = [{
                "Placa": "ABC-123",
                "Marca": "Volvo",
                "Modelo": "FH16",
                "Año": "2022",
                "Toneladas (TN)": "20.0",
                "Metraje (M3)": "15.0",
                "Tipo Vehículo": "Volquete",
                "Tipo Combustible": "Diesel",
                "Nombre Conductor": "Juan Pérez García",
                "Número Licencia": "Q12345678",
                "Revisión Técnica Emisión": "2024-01-10",
                "Revisión Técnica Vencimiento": "2025-01-10",
                "SOAT Inicio": "2024-02-01",
                "SOAT Fin": "2025-02-01",
                "MTC Número": "MTC-458712",
                "Extintor Vencimiento": "2025-06-30",
                "Cantidad Parihuelas": "12",
                "Días Alerta RT": "30",
                "Días Alerta SOAT": "30",
                "Días Alerta Extintor": "15",
                "Activo": "Sí",
                "Observaciones": "Vehículo ejemplo - puede eliminar esta fila"
            }]
            
            df = pd.DataFrame(template_data)
            
            # Crear Excel en memoria con formato
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Flota')
                
                workbook = writer.book
                worksheet = writer.sheets['Flota']
                
                # Estilo para encabezados
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajustar anchos de columna
                column_widths = {
                    'A': 12,  # Placa
                    'B': 15,  # Marca
                    'C': 15,  # Modelo
                    'D': 8,   # Año
                    'E': 16,  # Toneladas
                    'F': 14,  # Metraje
                    'G': 18,  # Tipo Vehículo
                    'H': 18,  # Tipo Combustible
                    'I': 20,  # Nombre Conductor
                    'J': 18,  # Número Licencia
                    'K': 25,  # RT Emisión
                    'L': 28,  # RT Vencimiento
                    'M': 14,  # SOAT Inicio
                    'N': 14,  # SOAT Fin
                    'O': 16,  # MTC Número
                    'P': 22,  # Extintor Vencimiento
                    'Q': 20,  # Cantidad Parihuelas
                    'R': 16,  # Días Alerta RT
                    'S': 18,  # Días Alerta SOAT
                    'T': 22,  # Días Alerta Extintor
                    'U': 10,  # Activo
                    'V': 40   # Observaciones
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Agregar instrucciones en una hoja separada
                instructions_data = {
                    "Campo": [
                        "Placa",
                        "Marca",
                        "Modelo",
                        "Año",
                        "Toneladas (TN)",
                        "Metraje (M3)",
                        "Tipo Vehículo",
                        "Tipo Combustible",
                        "Nombre Conductor",
                        "Número Licencia",
                        "Revisión Técnica Emisión",
                        "Revisión Técnica Vencimiento",
                        "SOAT Inicio",
                        "SOAT Fin",
                        "MTC Número",
                        "Extintor Vencimiento",
                        "Cantidad Parihuelas",
                        "Días Alerta RT",
                        "Días Alerta SOAT",
                        "Días Alerta Extintor",
                        "Activo",
                        "Observaciones"
                    ],
                    "Obligatorio": [
                        "SÍ", "SÍ", "SÍ", "SÍ", "SÍ", "SÍ", "SÍ",
                        "NO", "NO", "NO", "NO", "NO", "NO", "NO",
                        "NO", "NO", "NO", "NO", "NO", "NO", "NO", "NO"
                    ],
                    "Descripción": [
                        "Placa del vehículo (único)",
                        "Marca del vehículo",
                        "Modelo del vehículo",
                        "Año de fabricación (entre 1990 y 2025)",
                        "Capacidad de carga en toneladas",
                        "Capacidad volumétrica en metros cúbicos",
                        "Tipo: Volquete, Furgón, Plataforma, Tanque, Cisterna, etc.",
                        "Tipo de combustible: Diesel, Gasolina, GNV, Eléctrico",
                        "Nombre completo del conductor (opcional)",
                        "Número de licencia de conducir (único, opcional)",
                        "Fecha de emisión de la revisión técnica (formato: YYYY-MM-DD)",
                        "Fecha de vencimiento de la revisión técnica (formato: YYYY-MM-DD)",
                        "Fecha de inicio de vigencia del SOAT (formato: YYYY-MM-DD)",
                        "Fecha de fin de vigencia del SOAT (formato: YYYY-MM-DD)",
                        "Número de autorización o registro MTC",
                        "Fecha de vencimiento del extintor (formato: YYYY-MM-DD)",
                        "Cantidad de parihuelas que transporta",
                        "Días de anticipación para alertar vencimiento RT (por defecto: 30)",
                        "Días de anticipación para alertar vencimiento SOAT (por defecto: 30)",
                        "Días de anticipación para alertar vencimiento extintor (por defecto: 15)",
                        "Estado del vehículo: Sí o No (por defecto: Sí)",
                        "Observaciones o notas adicionales"
                    ],
                    "Ejemplo": [
                        "ABC-123",
                        "Volvo",
                        "FH16",
                        "2022",
                        "20.0",
                        "15.0",
                        "Volquete",
                        "Diesel",
                        "Juan Pérez García",
                        "Q12345678",
                        "2024-01-10",
                        "2025-01-10",
                        "2024-02-01",
                        "2025-02-01",
                        "MTC-458712",
                        "2025-06-30",
                        "12",
                        "30",
                        "30",
                        "15",
                        "Sí",
                        "Unidad operativa"
                    ]
                }
                
                df_instructions = pd.DataFrame(instructions_data)
                df_instructions.to_excel(writer, sheet_name='Instrucciones', index=False)
                
                # Formatear hoja de instrucciones
                ws_instructions = writer.sheets['Instrucciones']
                
                for cell in ws_instructions[1]:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                ws_instructions.column_dimensions['A'].width = 30
                ws_instructions.column_dimensions['B'].width = 15
                ws_instructions.column_dimensions['C'].width = 60
                ws_instructions.column_dimensions['D'].width = 30
                
                # Ajustar altura de filas en instrucciones
                for row in ws_instructions.iter_rows(min_row=2, max_row=ws_instructions.max_row):
                    ws_instructions.row_dimensions[row[0].row].height = 30
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al generar plantilla Excel: {str(e)}")
            raise
    
    def get_flotas_con_documentos_vencidos(self, dias_anticipacion: int = 30) -> Dict[str, List[dict]]:
        """Obtener vehículos con documentos vencidos o por vencer"""
        try:
            from datetime import timedelta
            hoy = datetime.now()
            fecha_limite = hoy + timedelta(days=dias_anticipacion)
            
            # Query base
            base_query = {"activo": True}
            
            # Revisión técnica vencida o por vencer
            revision_tecnica_query = {
                **base_query,
                "revision_tecnica_vencimiento": {
                    "$lte": fecha_limite,
                    "$ne": None
                }
            }
            revision_tecnica = list(self.collection.find(revision_tecnica_query))
            
            # SOAT vencido o por vencer
            soat_query = {
                **base_query,
                "soat_vigencia_fin": {
                    "$lte": fecha_limite,
                    "$ne": None
                }
            }
            soat = list(self.collection.find(soat_query))
            
            # Extintor vencido o por vencer
            extintor_query = {
                **base_query,
                "extintor_vencimiento": {
                    "$lte": fecha_limite,
                    "$ne": None
                }
            }
            extintor = list(self.collection.find(extintor_query))
            
            # Convertir y preparar respuestas
            return {
                "revision_tecnica": [self._prepare_flota_response(flota) for flota in revision_tecnica],
                "soat": [self._prepare_flota_response(flota) for flota in soat],
                "extintor": [self._prepare_flota_response(flota) for flota in extintor]
            }
            
        except Exception as e:
            logger.error(f"Error al obtener vehículos con documentos vencidos: {str(e)}")
            return {"revision_tecnica": [], "soat": [], "extintor": []}

def safe_regex(value: str):
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    return {"$regex": re.escape(value), "$options": "i"}