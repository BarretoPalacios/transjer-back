from typing import List, Optional, Dict, Any
from bson import ObjectId
from app.modules.utils.core.code_generator.code_generator import generate_sequential_code
from app.core.database import get_database
from app.modules.dataservice.models.personal import Personal
from app.modules.dataservice.schemas.personal_schema import PersonalCreate, PersonalUpdate, PersonalFilter
from datetime import datetime, date, timedelta
import pandas as pd
from io import BytesIO
import logging
import json
from math import ceil
import re

logger = logging.getLogger(__name__)

class PersonalService:
    def __init__(self, db):
        self.db = db 
        self.collection = db["personal"]
    
    def create_personal(self, personal_data: dict) -> dict:
        try:
            # 1️⃣ Verificar DNI (regla de negocio)
            existing_doc = self.collection.find_one({
                "dni": personal_data["dni"]
            })
            if existing_doc:
                raise ValueError(
                    f"El DNI {personal_data['dni']} ya está registrado"
                )

            # 2️⃣ Convertir objetos date a datetime para MongoDB
            personal_data = self._convert_dates_to_datetime(personal_data)

            # 3️⃣ Generar código automáticamente
            codigo_personal = generate_sequential_code(
                counters_collection=self.db["counters"],
                target_collection=self.collection,
                sequence_name="personal",
                field_name="codigo_personal",
                prefix="PER-",
                length=10
            )

            personal_data["codigo_personal"] = codigo_personal

            # 4️⃣ Crear modelo
            personal_model = Personal(**personal_data)

            # 5️⃣ Insertar
            result = self.collection.insert_one(
                personal_model.model_dump(by_alias=True)
            )

            # 6️⃣ Retornar creado
            created_personal = self.collection.find_one(
                {"_id": result.inserted_id}
            )
            created_personal["id"] = str(created_personal["_id"])
            del created_personal["_id"]
            
            # Convertir fechas de vuelta a date para la respuesta
            created_personal = self._convert_dates_to_date(created_personal)

            return created_personal

        except Exception as e:
            logger.error(f"Error al crear personal: {str(e)}")
            raise
    
    def get_personal_by_id(self, personal_id: str) -> Optional[dict]:
        """Obtener personal por ID"""
        try:
            if not ObjectId.is_valid(personal_id):
                return None
            
            personal = self.collection.find_one({"_id": ObjectId(personal_id)})
            if personal:
                personal["id"] = str(personal["_id"])
                del personal["_id"]
                # Convertir fechas de vuelta a date para la respuesta
                personal = self._convert_dates_to_date(personal)
            return personal
            
        except Exception as e:
            logger.error(f"Error al obtener personal: {str(e)}")
            return None
    
    def get_personal_by_codigo(self, codigo_personal: str) -> Optional[dict]:
        """Obtener personal por código"""
        try:
            personal = self.collection.find_one({"codigo_personal": codigo_personal})
            if personal:
                personal["id"] = str(personal["_id"])
                del personal["_id"]
                # Convertir fechas de vuelta a date para la respuesta
                personal = self._convert_dates_to_date(personal)
            return personal
            
        except Exception as e:
            logger.error(f"Error al obtener personal por código: {str(e)}")
            return None
    
    def get_personal_by_dni(self, dni: str) -> Optional[dict]:
        """Obtener personal por DNI"""
        try:
            personal = self.collection.find_one({"dni": dni})
            if personal:
                personal["id"] = str(personal["_id"])
                del personal["_id"]
                # Convertir fechas de vuelta a date para la respuesta
                personal = self._convert_dates_to_date(personal)
            return personal
            
        except Exception as e:
            logger.error(f"Error al obtener personal por DNI: {str(e)}")
            return None
    
    def get_all_personal(
            self,
            filter_params: Optional[PersonalFilter] = None,
            page: int = 1,
            page_size: int = 10,
            sort_by: str = "fecha_registro",
            sort_order: str = "desc"
        ) -> dict:
            """Obtener todos los personal con filtros opcionales y paginación"""
            try:
                query = {}

                if filter_params:
                    # Filtros básicos
                    if filter_params.dni:
                        query["dni"] = safe_regex(filter_params.dni)

                    if filter_params.nombres_completos:
                        query["nombres_completos"] = safe_regex(filter_params.nombres_completos)

                    if filter_params.tipo:
                        query["tipo"] = filter_params.tipo

                    if filter_params.estado:
                        query["estado"] = filter_params.estado

                    if filter_params.licencia_conducir:
                        query["licencia_conducir"] = safe_regex(filter_params.licencia_conducir)

                    if filter_params.categoria_licencia:
                        query["categoria_licencia"] = filter_params.categoria_licencia

                    if filter_params.turno:
                        query["turno"] = filter_params.turno

                    if filter_params.banco:
                        query["banco"] = safe_regex(filter_params.banco)

                    if filter_params.telefono:
                        query["telefono"] = safe_regex(filter_params.telefono)

                    if filter_params.email:
                        query["email"] = safe_regex(filter_params.email)

                    if filter_params.contacto_emergencia:
                        query["contacto_emergencia"] = safe_regex(filter_params.contacto_emergencia)

                    # Filtros de rango de fechas
                    date_filters = {}
                    
                    if filter_params.fecha_ingreso_desde:
                        fecha_desde = filter_params.fecha_ingreso_desde
                        if isinstance(fecha_desde, date):
                            fecha_desde = datetime.combine(fecha_desde, datetime.min.time())
                        date_filters["$gte"] = fecha_desde
                    
                    if filter_params.fecha_ingreso_hasta:
                        fecha_hasta = filter_params.fecha_ingreso_hasta
                        if isinstance(fecha_hasta, date):
                            fecha_hasta = datetime.combine(fecha_hasta, datetime.min.time())
                        date_filters["$lte"] = fecha_hasta
                    
                    if date_filters:
                        query["fecha_ingreso"] = date_filters

                    # Filtros de rango de salario
                    salary_filters = {}
                    
                    if filter_params.salario_min is not None:
                        salary_filters["$gte"] = filter_params.salario_min
                    
                    if filter_params.salario_max is not None:
                        salary_filters["$lte"] = filter_params.salario_max
                    
                    if salary_filters:
                        query["salario"] = salary_filters

                # Limpiar filtros nulos
                query = {k: v for k, v in query.items() if v is not None}

                total = self.collection.count_documents(query)

                skip = (page - 1) * page_size

                # Ordenamiento
                sort_direction = -1 if sort_order == "desc" else 1
                sort_field = sort_by if sort_by in ["dni", "nombres_completos", "fecha_ingreso", "fecha_registro", "salario"] else "fecha_registro"

                personal_list = list(
                    self.collection.find(query)
                    .sort(sort_field, sort_direction)
                    .skip(skip)
                    .limit(page_size)
                )

                for personal in personal_list:
                    personal["id"] = str(personal["_id"])
                    del personal["_id"]
                    # Convertir fechas de vuelta a date para la respuesta
                    personal = self._convert_dates_to_date(personal)

                total_pages = ceil(total / page_size) if page_size > 0 else 0

                return {
                    "items": personal_list,
                    "total": total,
                    "page": page,
                    "page_size": page_size,
                    "total_pages": total_pages,
                    "has_next": page < total_pages,
                    "has_prev": page > 1
                }

            except Exception as e:
                logger.error(f"Error al obtener personal: {str(e)}")
                raise
    
    def update_personal(self, personal_id: str, update_data: dict) -> Optional[dict]:
        """Actualizar personal"""
        try:
            if not ObjectId.is_valid(personal_id):
                return None
            
            # Filtrar campos None
            update_dict = {k: v for k, v in update_data.items() if v is not None}
            
            if not update_dict:
                return self.get_personal_by_id(personal_id)
            
            # Convertir objetos date a datetime para MongoDB
            update_dict = self._convert_dates_to_datetime(update_dict)
            
            # Si se actualiza el DNI, verificar que no exista otro personal con el mismo
            if "dni" in update_dict:
                dni_nuevo = update_dict["dni"]
                
                existing = self.collection.find_one({
                    "dni": dni_nuevo,
                    "_id": {"$ne": ObjectId(personal_id)}
                }) 
                
                if existing:
                    raise ValueError(f"Ya existe personal con el DNI {dni_nuevo}")
            
            # Actualizar en base de datos
            self.collection.update_one(
                {"_id": ObjectId(personal_id)},
                {"$set": update_dict}
            )
            
            return self.get_personal_by_id(personal_id)
            
        except Exception as e:
            logger.error(f"Error al actualizar personal: {str(e)}")
            raise
    
    def delete_personal(self, personal_id: str) -> bool:
        """Eliminar personal"""
        try:
            if not ObjectId.is_valid(personal_id):
                return False
            
            result = self.collection.delete_one({"_id": ObjectId(personal_id)})
            return result.deleted_count > 0
            
        except Exception as e:
            logger.error(f"Error al eliminar personal: {str(e)}")
            return False
    
    def export_to_excel(self, filter_params: Optional[PersonalFilter] = None) -> BytesIO:
        """Exportar personal a Excel"""
        try:
            personal_list = self.get_all_personal_sin_paginacion(filter_params)
            
            if not personal_list:
                # Crear DataFrame vacío
                df = pd.DataFrame(columns=[
                    "ID", "Código", "DNI", "Nombres Completos", "Tipo", "Estado",
                    "Fecha Ingreso", "Fecha Nacimiento", "Teléfono", "Email",
                    "Dirección", "Licencia Conducir", "Categoría Licencia",
                    "Fecha Venc. Licencia", "Turno", "Salario", "Banco",
                    "Número Cuenta", "Contacto Emergencia", "Teléfono Emergencia",
                    "Observaciones", "Fecha Registro"
                ])
            else:
                # Preparar datos para Excel
                excel_data = []
                for personal in personal_list:
                    excel_data.append({
                        "ID": personal.get("id", ""),
                        "Código": personal.get("codigo_personal", ""),
                        "DNI": personal.get("dni", ""),
                        "Nombres Completos": personal.get("nombres_completos", ""),
                        "Tipo": personal.get("tipo", ""),
                        "Estado": personal.get("estado", ""),
                        "Fecha Ingreso": personal.get("fecha_ingreso", "").strftime("%Y-%m-%d") if personal.get("fecha_ingreso") else "",
                        "Fecha Nacimiento": personal.get("fecha_nacimiento", "").strftime("%Y-%m-%d") if personal.get("fecha_nacimiento") else "",
                        "Teléfono": personal.get("telefono", ""),
                        "Email": personal.get("email", ""),
                        "Dirección": personal.get("direccion", ""),
                        "Licencia Conducir": personal.get("licencia_conducir", ""),
                        "Categoría Licencia": personal.get("categoria_licencia", ""),
                        "Fecha Venc. Licencia": personal.get("fecha_venc_licencia", "").strftime("%Y-%m-%d") if personal.get("fecha_venc_licencia") else "",
                        "Turno": personal.get("turno", ""),
                        "Salario": personal.get("salario", ""),
                        "Banco": personal.get("banco", ""),
                        "Número Cuenta": personal.get("numero_cuenta", ""),
                        "Contacto Emergencia": personal.get("contacto_emergencia", ""),
                        "Teléfono Emergencia": personal.get("telefono_emergencia", ""),
                        "Observaciones": personal.get("observaciones", ""),
                        "Fecha Registro": personal.get("fecha_registro", "").strftime("%Y-%m-%d %H:%M:%S") if personal.get("fecha_registro") else ""
                    })
                
                df = pd.DataFrame(excel_data)
            
            # Crear Excel en memoria
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Personal')
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {str(e)}")
            raise
    
    def import_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Importar personal desde Excel"""
        try:
            import io
            
            df = pd.read_excel(io.BytesIO(file_content))
            
            created = 0
            updated = 0
            errors = []
            skipped = 0
            
            for index, row in df.iterrows():
                try:
                    # Construir datos del personal (campos obligatorios)
                    personal_data = {
                        "dni": str(row.get("DNI", "")).strip(),
                        "nombres_completos": str(row.get("Nombres Completos", "")).strip(),
                        "tipo": str(row.get("Tipo", "")).strip(),
                    }
                    
                    # Campos opcionales
                    campos_opcionales = {
                        "estado": "Estado",
                        "fecha_ingreso": "Fecha Ingreso",
                        "fecha_nacimiento": "Fecha Nacimiento",
                        "telefono": "Teléfono",
                        "email": "Email",
                        "direccion": "Dirección",
                        "licencia_conducir": "Licencia Conducir",
                        "categoria_licencia": "Categoría Licencia",
                        "fecha_venc_licencia": "Fecha Venc. Licencia",
                        "turno": "Turno",
                        "salario": "Salario",
                        "banco": "Banco",
                        "numero_cuenta": "Número Cuenta",
                        "contacto_emergencia": "Contacto Emergencia",
                        "telefono_emergencia": "Teléfono Emergencia",
                        "observaciones": "Observaciones"
                    }
                    
                    for campo_db, campo_excel in campos_opcionales.items():
                        valor = row.get(campo_excel)
                        if pd.notna(valor) and str(valor).strip() and str(valor).strip().lower() != "nan":
                            valor_limpio = str(valor).strip()
                            
                            # Convertir campos especiales
                            if campo_db in ["salario"]:
                                try:
                                    personal_data[campo_db] = float(valor_limpio)
                                except (ValueError, TypeError):
                                    personal_data[campo_db] = 0.0
                            elif campo_db in ["fecha_ingreso", "fecha_nacimiento", "fecha_venc_licencia"]:
                                try:
                                    # Intentar parsear fecha
                                    if isinstance(valor_limpio, (datetime, date)):
                                        # Ya es una fecha
                                        personal_data[campo_db] = valor_limpio
                                    else:
                                        # Intentar parsear string a datetime
                                        personal_data[campo_db] = datetime.strptime(valor_limpio, "%Y-%m-%d")
                                except Exception as parse_error:
                                    # Si no se puede parsear, intentar con otros formatos
                                    try:
                                        personal_data[campo_db] = datetime.strptime(valor_limpio, "%d/%m/%Y")
                                    except:
                                        errors.append(f"Fila {index + 2}: Formato de fecha inválido para {campo_excel}: {valor_limpio}")
                                        continue
                            else:
                                personal_data[campo_db] = valor_limpio
                    
                    # Validar campos obligatorios
                    if not personal_data.get("dni"):
                        errors.append(f"Fila {index + 2}: DNI es requerido")
                        continue
                    
                    if not personal_data.get("nombres_completos"):
                        errors.append(f"Fila {index + 2}: Nombres completos son requeridos")
                        continue
                    
                    if not personal_data.get("tipo"):
                        errors.append(f"Fila {index + 2}: Tipo de personal es requerido")
                        continue
                    
                    # Validar DNI (solo números)
                    if not re.match(r'^\d+$', personal_data["dni"]):
                        errors.append(f"Fila {index + 2}: DNI debe contener solo números")
                        continue
                    
                    # Validar tipo de personal
                    tipos_permitidos = ['Conductor', 'Auxiliar', 'Operario', 'Administrativo', 
                                       'Supervisor', 'Mecánico', 'Almacenero']
                    if personal_data["tipo"] not in tipos_permitidos:
                        errors.append(f"Fila {index + 2}: Tipo debe ser uno de: {', '.join(tipos_permitidos)}")
                        continue
                    
                    # Estado por defecto
                    if not personal_data.get("estado"):
                        personal_data["estado"] = "Activo"
                    
                    # Verificar si ya existe por código (si viene en el Excel)
                    codigo_excel = str(row.get("Código", "")).strip()
                    if codigo_excel and codigo_excel not in ["", "nan", "None"]:
                        existing = self.collection.find_one({"codigo_personal": codigo_excel})
                        if existing:
                            # Actualizar personal existente
                            personal_data.pop("fecha_registro", None)  # No actualizar fecha de registro
                            # Convertir fechas a datetime
                            personal_data = self._convert_dates_to_datetime(personal_data)
                            
                            self.collection.update_one(
                                {"codigo_personal": codigo_excel},
                                {"$set": personal_data}
                            )
                            updated += 1
                            continue
                    
                    # Verificar si existe por DNI
                    existing_dni = self.collection.find_one({
                        "dni": personal_data["dni"]
                    })
                    
                    if existing_dni:
                        # Personal duplicado, lo saltamos
                        skipped += 1
                        errors.append(f"Fila {index + 2}: Personal duplicado - DNI {personal_data['dni']} ya existe")
                        continue
                    
                    # Crear nuevo personal
                    self.create_personal(personal_data)
                    created += 1
                        
                except Exception as e:
                    errors.append(f"Fila {index + 2}: {str(e)}")
                    continue
            
            return {
                "total_rows": len(df),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "has_errors": len(errors) > 0,
                "success_rate": f"{((created + updated) / len(df) * 100):.1f}%" if len(df) > 0 else "0%"
            }
            
        except Exception as e:
            logger.error(f"Error al importar desde Excel: {str(e)}")
            raise

    def get_stats(self) -> Dict[str, Any]:
        """Obtener estadísticas de personal"""
        try:
            total = self.collection.count_documents({})
            activos = self.collection.count_documents({"estado": "Activo"})
            inactivos = self.collection.count_documents({"estado": "Inactivo"})
            licencia = self.collection.count_documents({"estado": "Licencia"})
            vacaciones = self.collection.count_documents({"estado": "Vacaciones"})
            
            # Agrupar por tipo de personal
            pipeline_tipo = [
                {"$group": {"_id": "$tipo", "count": {"$sum": 1}}}
            ]
            
            tipos = {}
            for result in self.collection.aggregate(pipeline_tipo):
                tipos[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Agrupar por estado
            pipeline_estado = [
                {"$group": {"_id": "$estado", "count": {"$sum": 1}}}
            ]
            
            estados = {}
            for result in self.collection.aggregate(pipeline_estado):
                estados[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Agrupar por turno
            pipeline_turno = [
                {"$group": {"_id": "$turno", "count": {"$sum": 1}}}
            ]
            
            turnos = {}
            for result in self.collection.aggregate(pipeline_turno):
                turnos[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Calcular promedio de salario
            pipeline_salario = [
                {"$match": {"salario": {"$exists": True, "$ne": None}}},
                {"$group": {"_id": None, "promedio": {"$avg": "$salario"}}}
            ]
            
            salario_promedio = 0
            for result in self.collection.aggregate(pipeline_salario):
                salario_promedio = result.get("promedio", 0)
            
            # Contar licencias por vencer (próximos 30 días)
            hoy = datetime.now()
            fecha_limite = hoy + timedelta(days=30)
            
            licencias_por_vencer = self.collection.count_documents({
                "fecha_venc_licencia": {
                    "$gte": hoy,
                    "$lte": fecha_limite
                }
            })
            
            # Contar personal reciente (últimos 30 días)
            fecha_reciente = hoy - timedelta(days=30)
            personal_reciente = self.collection.count_documents({
                "fecha_ingreso": {
                    "$gte": fecha_reciente
                }
            })
            
            return {
                "total_personal": total,
                "total_activos": activos,
                "total_inactivos": inactivos,
                "por_tipo": tipos,
                "por_estado": estados,
                "por_turno": turnos,
                "promedio_salario": round(salario_promedio, 2) if salario_promedio else 0,
                "licencias_por_vencer": licencias_por_vencer,
                "personal_reciente": personal_reciente
            }
            
        except Exception as e:
            logger.error(f"Error al obtener estadísticas: {str(e)}")
            return {}

    def get_all_personal_sin_paginacion(
            self,
            filter_params: Optional[PersonalFilter] = None
        ) -> List[dict]:
            """Obtener TODO el personal sin paginación (para exportación)"""
            try:
                query = {}

                if filter_params:
                    # Filtros básicos (igual que en get_all_personal)
                    if filter_params.dni:
                        query["dni"] = safe_regex(filter_params.dni)

                    if filter_params.nombres_completos:
                        query["nombres_completos"] = safe_regex(filter_params.nombres_completos)

                    if filter_params.tipo:
                        query["tipo"] = filter_params.tipo

                    if filter_params.estado:
                        query["estado"] = filter_params.estado

                    if filter_params.licencia_conducir:
                        query["licencia_conducir"] = safe_regex(filter_params.licencia_conducir)

                    if filter_params.categoria_licencia:
                        query["categoria_licencia"] = filter_params.categoria_licencia

                    if filter_params.turno:
                        query["turno"] = filter_params.turno

                # Limpiar filtros None
                query = {k: v for k, v in query.items() if v is not None}

                personal_list = list(
                    self.collection
                    .find(query)
                    .sort("nombres_completos", 1)
                )

                for personal in personal_list:
                    personal["id"] = str(personal["_id"])
                    del personal["_id"]
                    # Convertir fechas de vuelta a date para la respuesta
                    personal = self._convert_dates_to_date(personal)

                return personal_list

            except Exception as e:
                logger.error(f"Error al obtener personal (sin paginación): {str(e)}")
                raise

    def generate_excel_template(self) -> BytesIO:
        """Generar plantilla de Excel vacía para importación"""
        try:
            # Fecha de ejemplo
            fecha_ejemplo = datetime.now().date()
            
            # Crear DataFrame con columnas y una fila de ejemplo
            template_data = [{
                "DNI": "87654321",
                "Nombres Completos": "JUAN CARLOS PEREZ GARCIA",
                "Tipo": "Conductor",
                "Estado": "Activo",
                "Fecha Ingreso": fecha_ejemplo.strftime("%Y-%m-%d"),
                "Fecha Nacimiento": "1985-07-20",
                "Teléfono": "987654321",
                "Email": "juan.perez@empresa.com",
                "Dirección": "Jr. Los Olivos 456, Lima",
                "Licencia Conducir": "Q12345678",
                "Categoría Licencia": "A-III-b",
                "Fecha Venc. Licencia": (fecha_ejemplo + timedelta(days=365)).strftime("%Y-%m-%d"),
                "Turno": "Día",
                "Salario": 2500.00,
                "Banco": "Banco de Crédito del Perú",
                "Número Cuenta": "19312345678901",
                "Contacto Emergencia": "María Pérez",
                "Teléfono Emergencia": "965432187",
                "Observaciones": "Conductor responsable - puede eliminar esta fila"
            }]
            
            df = pd.DataFrame(template_data)
            
            # Crear Excel en memoria con formato
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Personal')
                
                workbook = writer.book
                worksheet = writer.sheets['Personal']
                
                # Estilo para encabezados
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajustar anchos de columna
                column_widths = {
                    'A': 15,  # DNI
                    'B': 35,  # Nombres Completos
                    'C': 20,  # Tipo
                    'D': 15,  # Estado
                    'E': 15,  # Fecha Ingreso
                    'F': 15,  # Fecha Nacimiento
                    'G': 15,  # Teléfono
                    'H': 30,  # Email
                    'I': 40,  # Dirección
                    'J': 20,  # Licencia Conducir
                    'K': 20,  # Categoría Licencia
                    'L': 20,  # Fecha Venc. Licencia
                    'M': 15,  # Turno
                    'N': 15,  # Salario
                    'O': 25,  # Banco
                    'P': 25,  # Número Cuenta
                    'Q': 25,  # Contacto Emergencia
                    'R': 20,  # Teléfono Emergencia
                    'S': 40   # Observaciones
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Agregar instrucciones en una hoja separada
                instructions_data = {
                    "Campo": [
                        "DNI",
                        "Nombres Completos",
                        "Tipo",
                        "Estado",
                        "Fecha Ingreso",
                        "Fecha Nacimiento",
                        "Teléfono",
                        "Email",
                        "Dirección",
                        "Licencia Conducir",
                        "Categoría Licencia",
                        "Fecha Venc. Licencia",
                        "Turno",
                        "Salario",
                        "Banco",
                        "Número Cuenta",
                        "Contacto Emergencia",
                        "Teléfono Emergencia",
                        "Observaciones"
                    ],
                    "Obligatorio": [
                        "SÍ", "SÍ", "SÍ", "NO",
                        "NO", "NO", "NO", "NO", "NO",
                        "NO", "NO", "NO", "NO", "NO",
                        "NO", "NO", "NO", "NO", "NO"
                    ],
                    "Descripción": [
                        "Documento Nacional de Identidad (solo números, 8-15 dígitos)",
                        "Nombres y apellidos completos",
                        "Tipo de personal: Conductor, Auxiliar, Operario, Administrativo, Supervisor, Mecánico, Almacenero",
                        "Estado: Activo, Inactivo, Licencia, Vacaciones (por defecto: Activo)",
                        "Fecha de ingreso a la empresa (formato: YYYY-MM-DD)",
                        "Fecha de nacimiento (formato: YYYY-MM-DD)",
                        "Número de teléfono",
                        "Correo electrónico",
                        "Dirección completa",
                        "Número de licencia de conducir (solo para conductores)",
                        "Categoría de licencia: A-I, A-II-a, A-II-b, A-III-a, A-III-b, A-III-c",
                        "Fecha de vencimiento de licencia (formato: YYYY-MM-DD)",
                        "Turno de trabajo: Día, Noche, Rotativo",
                        "Salario mensual (número)",
                        "Banco para pago de salario",
                        "Número de cuenta bancaria",
                        "Nombre de contacto de emergencia",
                        "Teléfono de contacto de emergencia",
                        "Observaciones o notas adicionales"
                    ],
                    "Ejemplo": [
                        "87654321",
                        "JUAN CARLOS PEREZ GARCIA",
                        "Conductor",
                        "Activo",
                        "2023-01-15",
                        "1985-07-20",
                        "987654321",
                        "juan.perez@empresa.com",
                        "Jr. Los Olivos 456, Lima",
                        "Q12345678",
                        "A-III-b",
                        "2025-12-31",
                        "Día",
                        "2500.00",
                        "Banco de Crédito del Perú",
                        "19312345678901",
                        "María Pérez",
                        "965432187",
                        "Conductor con experiencia en ruta"
                    ]
                }
                
                df_instructions = pd.DataFrame(instructions_data)
                df_instructions.to_excel(writer, sheet_name='Instrucciones', index=False)
                
                # Formatear hoja de instrucciones
                ws_instructions = writer.sheets['Instrucciones']
                
                for cell in ws_instructions[1]:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                ws_instructions.column_dimensions['A'].width = 25
                ws_instructions.column_dimensions['B'].width = 15
                ws_instructions.column_dimensions['C'].width = 60
                ws_instructions.column_dimensions['D'].width = 30
                
                # Ajustar altura de filas en instrucciones
                for row in ws_instructions.iter_rows(min_row=2, max_row=ws_instructions.max_row):
                    ws_instructions.row_dimensions[row[0].row].height = 30
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al generar plantilla Excel: {str(e)}")
            raise

    def bulk_update_status(self, personal_ids: List[str], nuevo_estado: str, motivo: str = None, fecha_efectiva: date = None) -> Dict[str, Any]:
        """Actualizar estado de múltiples personal"""
        try:
            # Convertir IDs a ObjectId
            object_ids = []
            for pid in personal_ids:
                if ObjectId.is_valid(pid):
                    object_ids.append(ObjectId(pid))
            
            if not object_ids:
                return {"updated": 0, "errors": ["No hay IDs válidos"]}
            
            # Construir update data
            update_data = {"estado": nuevo_estado}
            
            if motivo:
                update_data["motivo_cambio_estado"] = motivo
            
            if fecha_efectiva:
                # Convertir date a datetime
                update_data["fecha_efectiva_cambio"] = datetime.combine(fecha_efectiva, datetime.min.time())
            
            update_data["fecha_ultima_modificacion"] = datetime.now()
            
            # Realizar actualización masiva
            result = self.collection.update_many(
                {"_id": {"$in": object_ids}},
                {"$set": update_data}
            )
            
            return {
                "updated": result.modified_count,
                "matched": result.matched_count,
                "errors": []
            }
            
        except Exception as e:
            logger.error(f"Error en actualización masiva: {str(e)}")
            return {"updated": 0, "errors": [str(e)]}

    # ===== MÉTODOS AUXILIARES PARA MANEJO DE FECHAS =====
    
    def _convert_dates_to_datetime(self, data_dict: dict) -> dict:
        """Convertir objetos date a datetime para MongoDB"""
        date_fields = ['fecha_ingreso', 'fecha_nacimiento', 'fecha_venc_licencia', 'fecha_registro']
        
        result = data_dict.copy()
        for field in date_fields:
            if field in result and isinstance(result[field], date):
                # Convertir date a datetime (a medianoche del día específico)
                result[field] = datetime.combine(result[field], datetime.min.time())
        
        return result
    
    def _convert_dates_to_date(self, data_dict: dict) -> dict:
        """Convertir campos datetime a date para la respuesta JSON"""
        date_fields = ['fecha_ingreso', 'fecha_nacimiento', 'fecha_venc_licencia', 'fecha_registro']
        
        result = data_dict.copy()
        for field in date_fields:
            if field in result:
                if isinstance(result[field], datetime):
                    result[field] = result[field].date()
                elif isinstance(result[field], str):
                    # Si es string, intentar convertir
                    try:
                        dt = datetime.fromisoformat(result[field].replace('Z', '+00:00'))
                        result[field] = dt.date()
                    except:
                        # Mantener como está si no se puede convertir
                        pass
        
        return result

def safe_regex(value: str):
    """Función auxiliar para búsquedas seguras con regex"""
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    return {"$regex": re.escape(value), "$options": "i"}