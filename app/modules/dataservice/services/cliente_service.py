from typing import List, Optional, Dict, Any
from bson import ObjectId
from app.modules.utils.core.code_generator.code_generator import generate_sequential_code
from app.core.database import get_database
from app.modules.dataservice.models.cliente import Cliente
from app.modules.dataservice.schemas.cliente_schema import ClienteCreate, ClienteUpdate, ClienteFilter
from datetime import datetime
import pandas as pd
from io import BytesIO
import logging
import json
from math import ceil
import re

logger = logging.getLogger(__name__)

class ClienteService:
    def __init__(self, db):
        self.db = db 
        self.collection = db["clientes"]
    
    def create_cliente(self, cliente_data: dict) -> dict:
        try:
            # 1️⃣ Verificar documento (regla de negocio)
            existing_doc = self.collection.find_one({
                "tipo_documento": cliente_data["tipo_documento"],
                "numero_documento": cliente_data["numero_documento"]
            })
            if existing_doc:
                raise ValueError(
                    f"El {cliente_data['tipo_documento']} "
                    f"{cliente_data['numero_documento']} ya está registrado"
                )

            # 2️⃣ Generar código automáticamente
            codigo_cliente = generate_sequential_code(
                counters_collection=self.db["counters"],
                target_collection=self.collection,
                sequence_name="clientes",
                field_name="codigo_cliente",
                prefix="CLI-",
                length=10
            )

            cliente_data["codigo_cliente"] = codigo_cliente

            # 3️⃣ Crear modelo
            cliente_model = Cliente(**cliente_data)

            # 4️⃣ Insertar
            result = self.collection.insert_one(
                cliente_model.model_dump(by_alias=True)
            )

            # 5️⃣ Retornar creado
            created_cliente = self.collection.find_one(
                {"_id": result.inserted_id}
            )
            created_cliente["id"] = str(created_cliente["_id"])
            del created_cliente["_id"]

            return created_cliente

        except Exception as e:
            logger.error(f"Error al crear cliente: {str(e)}")
            raise
    
    def get_cliente_by_id(self, cliente_id: str) -> Optional[dict]:
        """Obtener cliente por ID"""
        try:
            if not ObjectId.is_valid(cliente_id):
                return None
            
            cliente = self.collection.find_one({"_id": ObjectId(cliente_id)})
            if cliente:
                cliente["id"] = str(cliente["_id"])
                del cliente["_id"]
            return cliente
            
        except Exception as e:
            logger.error(f"Error al obtener cliente: {str(e)}")
            return None
    
    def get_cliente_by_codigo(self, codigo_cliente: str) -> Optional[dict]:
        """Obtener cliente por código"""
        try:
            cliente = self.collection.find_one({"codigo_cliente": codigo_cliente})
            if cliente:
                cliente["id"] = str(cliente["_id"])
                del cliente["_id"]
            return cliente
            
        except Exception as e:
            logger.error(f"Error al obtener cliente por código: {str(e)}")
            return None
    
    def get_cliente_by_documento(self, tipo_documento: str, numero_documento: str) -> Optional[dict]:
        """Obtener cliente por tipo y número de documento"""
        try:
            cliente = self.collection.find_one({
                "tipo_documento": tipo_documento,
                "numero_documento": numero_documento
            })
            if cliente:
                cliente["id"] = str(cliente["_id"])
                del cliente["_id"]
            return cliente
            
        except Exception as e:
            logger.error(f"Error al obtener cliente por documento: {str(e)}")
            return None
    
    def get_all_clientes(
            self,
            filter_params: Optional[ClienteFilter] = None,
            page: int = 1,
            page_size: int = 10
        ) -> dict:
            """Obtener todos los clientes con filtros opcionales y paginación"""
            try:
                query = {}

                if filter_params:
                    if filter_params.codigo_cliente:
                        query["codigo_cliente"] = safe_regex(filter_params.codigo_cliente)

                    if filter_params.tipo_documento:
                        query["tipo_documento"] = filter_params.tipo_documento

                    if filter_params.numero_documento:
                        query["numero_documento"] = safe_regex(filter_params.numero_documento)

                    if filter_params.razon_social:
                        query["razon_social"] = safe_regex(filter_params.razon_social)

                    if filter_params.tipo_cliente:
                        query["tipo_cliente"] = filter_params.tipo_cliente

                    if filter_params.tipo_pago:
                        query["tipo_pago"] = filter_params.tipo_pago

                    if filter_params.dias_credito is not None:
                        query["dias_credito"] = filter_params.dias_credito

                    if filter_params.contacto_principal:
                        query["contacto_principal"] = safe_regex(filter_params.contacto_principal)

                    if filter_params.telefono:
                        query["telefono"] = safe_regex(filter_params.telefono)

                    if filter_params.estado:
                        query["estado"] = filter_params.estado

                    if filter_params.periodo_facturacion:
                        query["periodo_facturacion"] = filter_params.periodo_facturacion

                # Limpiar filtros nulos
                query = {k: v for k, v in query.items() if v is not None}

                total = self.collection.count_documents(query)

                skip = (page - 1) * page_size

                clientes = list(
                    self.collection.find(query)
                    .sort("razon_social", 1)
                    .skip(skip)
                    .limit(page_size)
                )

                for cliente in clientes:
                    cliente["id"] = str(cliente["_id"])
                    del cliente["_id"]

                total_pages = ceil(total / page_size) if page_size > 0 else 0

                return {
                    "items": clientes,
                    "total": total,
                    "page": page,
                    "page_size": page_size,
                    "total_pages": total_pages,
                    "has_next": page < total_pages,
                    "has_prev": page > 1
                }

            except Exception as e:
                logger.error(f"Error al obtener clientes: {str(e)}")
                raise
    
    def update_cliente(self, cliente_id: str, update_data: dict) -> Optional[dict]:
        """Actualizar cliente"""
        try:
            if not ObjectId.is_valid(cliente_id):
                return None
            
            # Filtrar campos None
            update_dict = {k: v for k, v in update_data.items() if v is not None}
            
            if not update_dict:
                return self.get_cliente_by_id(cliente_id)
            
            # Si se actualiza el documento, verificar que no exista otro cliente con el mismo
            if "numero_documento" in update_dict or "tipo_documento" in update_dict:
                cliente_actual = self.get_cliente_by_id(cliente_id)
                tipo_doc = update_dict.get("tipo_documento", cliente_actual.get("tipo_documento"))
                num_doc = update_dict.get("numero_documento", cliente_actual.get("numero_documento"))
                
                existing = self.collection.find_one({
                    "tipo_documento": tipo_doc,
                    "numero_documento": num_doc,
                    "_id": {"$ne": ObjectId(cliente_id)}
                }) 
                
                if existing:
                    raise ValueError(f"Ya existe un cliente con el {tipo_doc} {num_doc}")
            
            # Actualizar en base de datos
            self.collection.update_one(
                {"_id": ObjectId(cliente_id)},
                {"$set": update_dict}
            )
            
            return self.get_cliente_by_id(cliente_id)
            
        except Exception as e:
            logger.error(f"Error al actualizar cliente: {str(e)}")
            raise
    
    def delete_cliente(self, cliente_id: str) -> bool:
        """Eliminar cliente"""
        try:
            if not ObjectId.is_valid(cliente_id):
                return False
            
            result = self.collection.delete_one({"_id": ObjectId(cliente_id)})
            return result.deleted_count > 0
            
        except Exception as e:
            logger.error(f"Error al eliminar cliente: {str(e)}")
            return False
    
    def export_to_excel(self, filter_params: Optional[ClienteFilter] = None) -> BytesIO:
        """Exportar clientes a Excel"""
        try:
            clientes = self.get_all_clientes_sin_paginacion(filter_params)
            
            if not clientes:
                # Crear DataFrame vacío
                df = pd.DataFrame(columns=[
                    "ID", "Código Cliente", "Tipo Documento", "Número Documento",
                    "Razón Social", "Tipo Cliente", "Período Facturación",
                    "Tipo Pago", "Días Crédito",
                    "Contacto Principal", "Teléfono", "Email", "Dirección",
                    "Estado", "Website", "Fecha Registro", "Observaciones",
                    "Contactos", "Cuentas Facturación"
                ])
            else:
                # Preparar datos para Excel
                excel_data = []
                for cliente in clientes:
                    # Serializar contactos y cuentas a JSON string para Excel
                    contactos_str = ""
                    if cliente.get("contactos"):
                        contactos_str = json.dumps(cliente["contactos"], ensure_ascii=False)
                    
                    cuentas_str = ""
                    if cliente.get("cuentas_facturacion"):
                        cuentas_str = json.dumps(cliente["cuentas_facturacion"], ensure_ascii=False)
                    
                    excel_data.append({
                        "ID": cliente.get("id", ""),
                        "Código Cliente": cliente.get("codigo_cliente", ""),
                        "Tipo Documento": cliente.get("tipo_documento", ""),
                        "Número Documento": cliente.get("numero_documento", ""),
                        "Razón Social": cliente.get("razon_social", ""),
                        "Tipo Cliente": cliente.get("tipo_cliente", ""),
                        "Período Facturación": cliente.get("periodo_facturacion", ""),
                        "Tipo Pago": cliente.get("tipo_pago", ""),
                        "Días Crédito": cliente.get("dias_credito", 0),
                        "Contacto Principal": cliente.get("contacto_principal", ""),
                        "Teléfono": cliente.get("telefono", ""),
                        "Email": cliente.get("email", ""),
                        "Dirección": cliente.get("direccion", ""),
                        "Estado": cliente.get("estado", ""),
                        "Website": cliente.get("website", ""),
                        "Fecha Registro": cliente.get("fecha_registro", "").strftime("%Y-%m-%d %H:%M:%S") if cliente.get("fecha_registro") else "",
                        "Observaciones": cliente.get("observaciones", ""),
                        "Contactos": contactos_str,
                        "Cuentas Facturación": cuentas_str
                    })
                
                df = pd.DataFrame(excel_data)
            
            # Crear Excel en memoria
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Clientes')
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {str(e)}")
            raise
    
    def import_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Importar clientes desde Excel (formato simplificado)"""
        try:
            import io
            
            df = pd.read_excel(io.BytesIO(file_content))
            
            created = 0
            updated = 0
            errors = []
            skipped = 0
            
            for index, row in df.iterrows():
                try:
                    # Construir datos del cliente (solo campos simples)
                    cliente_data = {
                        "tipo_documento": str(row.get("Tipo Documento", "")).strip(),
                        "numero_documento": str(row.get("Número Documento", "")).strip(),
                        "razon_social": str(row.get("Razón Social", "")).strip(),
                        "estado": str(row.get("Estado", "activo")).strip(),
                        "fecha_registro": datetime.now()
                    }
                    
                    # Campos opcionales simples
                    campos_opcionales = {
                        "tipo_cliente": "Tipo Cliente",
                        "periodo_facturacion": "Período Facturación",
                        "periodo_facturacion_dias": "Período Facturación Días",
                        "tipo_pago": "Tipo Pago",
                        "dias_credito": "Días Crédito",
                        "contacto_principal": "Contacto Principal",
                        "telefono": "Teléfono",
                        "email": "Email",
                        "direccion": "Dirección",
                        "website": "Website",
                        "observaciones": "Observaciones"
                    }
                    
                    for campo_db, campo_excel in campos_opcionales.items():
                        valor = row.get(campo_excel)
                        if pd.notna(valor) and str(valor).strip() and str(valor).strip().lower() != "nan":
                            valor_limpio = str(valor).strip()
                            
                            # Convertir campos numéricos
                            if campo_db in ["dias_credito", "periodo_facturacion_dias"]:
                                try:
                                    cliente_data[campo_db] = int(float(valor_limpio))
                                except (ValueError, TypeError):
                                    cliente_data[campo_db] = 0
                            else:
                                cliente_data[campo_db] = valor_limpio
                    
                    # Validar campos obligatorios
                    if not cliente_data.get("tipo_documento"):
                        errors.append(f"Fila {index + 2}: Tipo de documento es requerido")
                        continue
                    
                    if not cliente_data.get("numero_documento"):
                        errors.append(f"Fila {index + 2}: Número de documento es requerido")
                        continue
                    
                    if not cliente_data.get("razon_social"):
                        errors.append(f"Fila {index + 2}: Razón social es requerida")
                        continue
                    
                    # Asegurar estado válido
                    if cliente_data.get("estado") not in ["activo", "inactivo", "suspendido"]:
                        cliente_data["estado"] = "activo"
                    
                    # Verificar si ya existe por código (si viene en el Excel)
                    codigo_excel = str(row.get("Código Cliente", "")).strip()
                    if codigo_excel and codigo_excel not in ["", "nan", "None"]:
                        existing = self.collection.find_one({"codigo_cliente": codigo_excel})
                        if existing:
                            # Actualizar cliente existente
                            cliente_data.pop("fecha_registro", None)  # No actualizar fecha de registro
                            self.collection.update_one(
                                {"codigo_cliente": codigo_excel},
                                {"$set": cliente_data}
                            )
                            updated += 1
                            continue
                    
                    # Verificar si existe por documento
                    existing_doc = self.collection.find_one({
                        "tipo_documento": cliente_data["tipo_documento"],
                        "numero_documento": cliente_data["numero_documento"]
                    })
                    
                    if existing_doc:
                        # Cliente duplicado, lo saltamos
                        skipped += 1
                        errors.append(f"Fila {index + 2}: Cliente duplicado - {cliente_data['tipo_documento']} {cliente_data['numero_documento']} ya existe")
                        continue
                    
                    # Crear nuevo cliente (sin código, se generará automáticamente)
                    self.create_cliente(cliente_data)
                    created += 1
                        
                except Exception as e:
                    errors.append(f"Fila {index + 2}: {str(e)}")
                    continue
            
            return {
                "total_rows": len(df),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "has_errors": len(errors) > 0,
                "success_rate": f"{((created + updated) / len(df) * 100):.1f}%" if len(df) > 0 else "0%"
            }
            
        except Exception as e:
            logger.error(f"Error al importar desde Excel: {str(e)}")
            raise

    def get_stats(self) -> Dict[str, Any]:
        """Obtener estadísticas de clientes"""
        try:
            total = self.collection.count_documents({})
            activos = self.collection.count_documents({"estado": "activo"})
            inactivos = self.collection.count_documents({"estado": "inactivo"})
            pendientes = self.collection.count_documents({"estado": "pendiente"})
            suspendidos = self.collection.count_documents({"estado": "suspendido"})
            
            # Agrupar por tipo de cliente (rubro comercial)
            pipeline_tipo = [
                {"$group": {"_id": "$tipo_cliente", "count": {"$sum": 1}}}
            ]
            
            tipos = {}
            for result in self.collection.aggregate(pipeline_tipo):
                tipos[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Agrupar por tipo de documento
            pipeline_doc = [
                {"$group": {"_id": "$tipo_documento", "count": {"$sum": 1}}}
            ]
            
            documentos = {}
            for result in self.collection.aggregate(pipeline_doc):
                documentos[result["_id"]] = result["count"]
            
            # Agrupar por tipo de pago
            pipeline_pago = [
                {"$group": {"_id": "$tipo_pago", "count": {"$sum": 1}}}
            ]
            
            pagos = {}
            for result in self.collection.aggregate(pipeline_pago):
                pagos[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Agrupar por período de facturación
            pipeline_periodo = [
                {"$group": {"_id": "$periodo_facturacion", "count": {"$sum": 1}}}
            ]
            
            periodos = {}
            for result in self.collection.aggregate(pipeline_periodo):
                periodos[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            return {
                "total": total,
                "activos": activos,
                "inactivos": inactivos,
                "pendientes": pendientes,
                "suspendidos": suspendidos,
                "por_tipo_cliente": tipos,
                "por_tipo_documento": documentos,
                "por_tipo_pago": pagos,
                "por_periodo_facturacion": periodos
            }
            
        except Exception as e:
            logger.error(f"Error al obtener estadísticas: {str(e)}")
            return {}

    def get_all_clientes_sin_paginacion(
            self,
            filter_params: Optional[ClienteFilter] = None
        ) -> List[dict]:
            """Obtener TODOS los clientes sin paginación (para exportación)"""
            try:
                query = {}

                if filter_params:
                    if filter_params.codigo_cliente:
                        query["codigo_cliente"] = safe_regex(filter_params.codigo_cliente)

                    if filter_params.tipo_documento:
                        query["tipo_documento"] = filter_params.tipo_documento

                    if filter_params.numero_documento:
                        query["numero_documento"] = safe_regex(filter_params.numero_documento)

                    if filter_params.razon_social:
                        query["razon_social"] = safe_regex(filter_params.razon_social)

                    if filter_params.tipo_cliente:
                        query["tipo_cliente"] = filter_params.tipo_cliente

                    if filter_params.tipo_pago:
                        query["tipo_pago"] = filter_params.tipo_pago

                    if filter_params.dias_credito is not None:
                        query["dias_credito"] = filter_params.dias_credito

                    if filter_params.contacto_principal:
                        query["contacto_principal"] = safe_regex(filter_params.contacto_principal)

                    if filter_params.telefono:
                        query["telefono"] = safe_regex(filter_params.telefono)

                    if filter_params.estado:
                        query["estado"] = filter_params.estado

                    if filter_params.periodo_facturacion:
                        query["periodo_facturacion"] = filter_params.periodo_facturacion

                # Limpiar filtros None
                query = {k: v for k, v in query.items() if v is not None}

                clientes = list(
                    self.collection
                    .find(query)
                    .sort("razon_social", 1)
                )

                for cliente in clientes:
                    cliente["id"] = str(cliente["_id"])
                    del cliente["_id"]

                return clientes

            except Exception as e:
                logger.error(f"Error al obtener clientes (sin paginación): {str(e)}")
                raise

    def generate_excel_template(self) -> BytesIO:
        """Generar plantilla de Excel vacía para importación"""
        try:
            # Crear DataFrame con columnas y una fila de ejemplo
            template_data = [{
                "Tipo Documento": "RUC",
                "Número Documento": "20123456789",
                "Razón Social": "EJEMPLO EMPRESA SAC",
                "Tipo Cliente": "mayorista",
                "Período Facturación": "Mensual",
                "Período Facturación Días": "30",
                "Tipo Pago": "Crédito",
                "Días Crédito": "30",
                "Contacto Principal": "Juan Pérez",
                "Teléfono": "987654321",
                "Email": "contacto@ejemplo.com",
                "Dirección": "Av. Ejemplo 123, Lima",
                "Website": "www.ejemplo.com",
                "Estado": "activo",
                "Observaciones": "Cliente ejemplo - puede eliminar esta fila"
            }]
            
            df = pd.DataFrame(template_data)
            
            # Crear Excel en memoria con formato
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Clientes')
                
                workbook = writer.book
                worksheet = writer.sheets['Clientes']
                
                # Estilo para encabezados
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajustar anchos de columna
                column_widths = {
                    'A': 18,  # Tipo Documento
                    'B': 18,  # Número Documento
                    'C': 35,  # Razón Social
                    'D': 18,  # Tipo Cliente
                    'E': 22,  # Período Facturación
                    'F': 25,  # Período Facturación Días
                    'G': 15,  # Tipo Pago
                    'H': 15,  # Días Crédito
                    'I': 25,  # Contacto Principal
                    'J': 15,  # Teléfono
                    'K': 30,  # Email
                    'L': 40,  # Dirección
                    'M': 25,  # Website
                    'N': 12,  # Estado
                    'O': 40   # Observaciones
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Agregar instrucciones en una hoja separada
                instructions_data = {
                    "Campo": [
                        "Tipo Documento",
                        "Número Documento",
                        "Razón Social",
                        "Estado",
                        "Tipo Cliente",
                        "Período Facturación",
                        "Período Facturación Días",
                        "Tipo Pago",
                        "Días Crédito",
                        "Contacto Principal",
                        "Teléfono",
                        "Email",
                        "Dirección",
                        "Website",
                        "Observaciones"
                    ],
                    "Obligatorio": [
                        "SÍ", "SÍ", "SÍ", "NO",
                        "NO", "NO", "NO", "NO", "NO",
                        "NO", "NO", "NO", "NO", "NO", "NO"
                    ],
                    "Descripción": [
                        "Tipo de documento: RUC, DNI, CE, PASAPORTE",
                        "Número del documento de identidad",
                        "Nombre o razón social del cliente",
                        "Estado del cliente: activo, inactivo, suspendido (por defecto: activo)",
                        "Tipo o rubro comercial del cliente",
                        "Período de facturación: Semanal, Quincenal, Mensual, Bimensual, etc.",
                        "Días del período de facturación (número)",
                        "Tipo de pago: Contado, Crédito, Crédito Factoring",
                        "Cantidad de días de crédito (número)",
                        "Nombre del contacto principal",
                        "Número de teléfono",
                        "Correo electrónico",
                        "Dirección completa del cliente",
                        "Sitio web",
                        "Observaciones o notas adicionales"
                    ],
                    "Ejemplo": [
                        "RUC",
                        "20123456789",
                        "EJEMPLO EMPRESA SAC",
                        "activo",
                        "mayorista",
                        "Mensual",
                        "30",
                        "Crédito",
                        "30",
                        "Juan Pérez",
                        "987654321",
                        "contacto@ejemplo.com",
                        "Av. Ejemplo 123, Lima",
                        "www.ejemplo.com",
                        "Cliente VIP"
                    ]
                }
                
                df_instructions = pd.DataFrame(instructions_data)
                df_instructions.to_excel(writer, sheet_name='Instrucciones', index=False)
                
                # Formatear hoja de instrucciones
                ws_instructions = writer.sheets['Instrucciones']
                
                for cell in ws_instructions[1]:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                ws_instructions.column_dimensions['A'].width = 30
                ws_instructions.column_dimensions['B'].width = 15
                ws_instructions.column_dimensions['C'].width = 60
                ws_instructions.column_dimensions['D'].width = 30
                
                # Ajustar altura de filas en instrucciones
                for row in ws_instructions.iter_rows(min_row=2, max_row=ws_instructions.max_row):
                    ws_instructions.row_dimensions[row[0].row].height = 30
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al generar plantilla Excel: {str(e)}")
            raise

    def get_cuentas_by_cliente_id(self, cliente_id: str) -> dict:
        """Obtener cuentas de facturación activas por ID de cliente"""
        try:
            cliente = self.get_cliente_by_id(cliente_id)
            
            if not cliente:
                return {}

            # Filtramos la lista de cuentas usando una comprensión de lista
            cuentas_totales = cliente.get("cuentas_facturacion", [])
            cuentas_activas = [c for c in cuentas_totales if c.get("estado") == "activa"]

            return {
                "cliente_id": cliente_id,
                "cuentas_facturacion": cuentas_activas
            }
        except Exception as e:
            logger.error(f"Error al obtener cuentas de facturación: {str(e)}")
            return {}

    def add_cuenta_to_cliente(self, cliente_id: str, cuenta_data: dict) -> dict:
        try:
            if not ObjectId.is_valid(cliente_id):
                raise ValueError("ID de cliente inválido")

            nueva_cuenta = {
                "nombre_cuenta": cuenta_data["nombre_cuenta"],
                "direccion_origen": cuenta_data.get("direccion_origen", ""),
                "tipo_pago": cuenta_data.get("tipo_pago", "Contado"),
                "dias_credito": int(cuenta_data.get("dias_credito", 0)),
                "limite_credito": float(cuenta_data.get("limite_credito", 0)),
                "estado": cuenta_data.get("estado", "activa"),
                "es_principal": cuenta_data.get("es_principal", False),
            }

            # 1️⃣ Asegurar que cuentas_facturacion sea un array
            self.collection.update_one(
                {
                    "_id": ObjectId(cliente_id),
                    "$or": [
                        {"cuentas_facturacion": None},
                        {"cuentas_facturacion": {"$exists": False}}
                    ]
                },
                {"$set": {"cuentas_facturacion": []}}
            )

            # 2️⃣ Si es principal, desactivar las demás
            if nueva_cuenta["es_principal"]:
                self.collection.update_one(
                    {"_id": ObjectId(cliente_id)},
                    {"$set": {"cuentas_facturacion.$[].es_principal": False}}
                )

            # 3️⃣ Push seguro
            result = self.collection.update_one(
                {"_id": ObjectId(cliente_id)},
                {"$push": {"cuentas_facturacion": nueva_cuenta}}
            )

            if result.matched_count == 0:
                raise ValueError("Cliente no encontrado")

            return nueva_cuenta

        except Exception as e:
            logger.error(f"Error al agregar cuenta de facturación: {str(e)}")
            raise


def safe_regex(value: str):
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    return {"$regex": re.escape(value), "$options": "i"}