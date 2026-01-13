from typing import List, Optional, Dict, Any
from bson import ObjectId
from app.modules.utils.core.code_generator.code_generator import generate_sequential_code
from app.core.database import get_database
from app.modules.dataservice.models.proveedor import Proveedor
from app.modules.dataservice.schemas.proveedor_schema import ProveedorCreate, ProveedorUpdate, ProveedorFilter
from datetime import datetime
import pandas as pd
from io import BytesIO
import logging
import json
from math import ceil
import re

logger = logging.getLogger(__name__)

class ProveedorService:
    def __init__(self, db):
        self.db = db 
        self.collection = db["proveedores"]
    
    def create_proveedor(self, proveedor_data: dict) -> dict:
        try:
            # 1️⃣ Verificar documento (regla de negocio)
            existing_doc = self.collection.find_one({
                "tipo_documento": proveedor_data["tipo_documento"],
                "numero_documento": proveedor_data["numero_documento"]
            })
            if existing_doc:
                raise ValueError(
                    f"El {proveedor_data['tipo_documento']} "
                    f"{proveedor_data['numero_documento']} ya está registrado"
                )

            # 2️⃣ Generar código automáticamente
            codigo_proveedor = generate_sequential_code(
                counters_collection=self.db["counters"],
                target_collection=self.collection,
                sequence_name="proveedores",
                field_name="codigo_proveedor",
                prefix="PROV-",
                length=10
            )

            proveedor_data["codigo_proveedor"] = codigo_proveedor

            # 3️⃣ Crear modelo
            proveedor_model = Proveedor(**proveedor_data)

            # 4️⃣ Insertar
            result = self.collection.insert_one(
                proveedor_model.model_dump(by_alias=True)
            )

            # 5️⃣ Retornar creado
            created_proveedor = self.collection.find_one(
                {"_id": result.inserted_id}
            )
            created_proveedor["id"] = str(created_proveedor["_id"])
            del created_proveedor["_id"]

            return created_proveedor

        except Exception as e:
            logger.error(f"Error al crear proveedor: {str(e)}")
            raise
    
    def get_proveedor_by_id(self, proveedor_id: str) -> Optional[dict]:
        """Obtener proveedor por ID"""
        try:
            if not ObjectId.is_valid(proveedor_id):
                return None
            
            proveedor = self.collection.find_one({"_id": ObjectId(proveedor_id)})
            if proveedor:
                proveedor["id"] = str(proveedor["_id"])
                del proveedor["_id"]
            return proveedor
            
        except Exception as e:
            logger.error(f"Error al obtener proveedor: {str(e)}")
            return None
    
    def get_proveedor_by_codigo(self, codigo_proveedor: str) -> Optional[dict]:
        """Obtener proveedor por código"""
        try:
            proveedor = self.collection.find_one({"codigo_proveedor": codigo_proveedor})
            if proveedor:
                proveedor["id"] = str(proveedor["_id"])
                del proveedor["_id"]
            return proveedor
            
        except Exception as e:
            logger.error(f"Error al obtener proveedor por código: {str(e)}")
            return None
    
    def get_proveedor_by_documento(self, tipo_documento: str, numero_documento: str) -> Optional[dict]:
        """Obtener proveedor por tipo y número de documento"""
        try:
            proveedor = self.collection.find_one({
                "tipo_documento": tipo_documento,
                "numero_documento": numero_documento
            })
            if proveedor:
                proveedor["id"] = str(proveedor["_id"])
                del proveedor["_id"]
            return proveedor
            
        except Exception as e:
            logger.error(f"Error al obtener proveedor por documento: {str(e)}")
            return None
    
    def get_all_proveedores(
        self, 
        filter_params: Optional[ProveedorFilter] = None,
        page: int = 1,
        page_size: int = 10
    ) -> dict:
        """Obtener todos los proveedores con filtros opcionales y paginación"""
        try:
            query = {}
            
            if filter_params:
                if filter_params.codigo_proveedor:
                    query["codigo_proveedor"] = safe_regex(filter_params.codigo_proveedor)
                
                if filter_params.tipo_documento:
                    query["tipo_documento"] = filter_params.tipo_documento
                
                if filter_params.numero_documento:
                    query["numero_documento"] = safe_regex(filter_params.numero_documento)
                
                if filter_params.razon_social:
                    query["razon_social"] = safe_regex(filter_params.razon_social)
                
                if filter_params.rubro_proveedor:
                    query["rubro_proveedor"] = filter_params.rubro_proveedor
                
                if filter_params.contacto_principal:
                    query["contacto_principal"] = safe_regex(filter_params.contacto_principal)
                
                if filter_params.telefono:
                    query["telefono"] = safe_regex(filter_params.telefono)
                
                if filter_params.estado:
                    query["estado"] = filter_params.estado
                
                if filter_params.servicio:
                    # Buscar en el array de servicios con safe_regex
                    query["servicios"] = safe_regex(filter_params.servicio)
            
            # Limpiar filtros nulos
            query = {k: v for k, v in query.items() if v is not None}
            
            # Contar total de documentos
            total = self.collection.count_documents(query)
            
            # Calcular skip
            skip = (page - 1) * page_size
            
            # Obtener proveedores paginados
            proveedores = list(
                self.collection.find(query)
                .sort("razon_social", 1)
                .skip(skip)
                .limit(page_size)
            )
            
            # Convertir ObjectId a string
            for proveedor in proveedores:
                proveedor["id"] = str(proveedor["_id"])
                del proveedor["_id"]
            
            # Calcular metadatos de paginación
            total_pages = ceil(total / page_size) if page_size > 0 else 0
            
            return {
                "items": proveedores,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": total_pages,
                "has_next": page < total_pages,
                "has_prev": page > 1
            }
            
        except Exception as e:
            logger.error(f"Error al obtener proveedores: {str(e)}")
            raise

    def update_proveedor(self, proveedor_id: str, update_data: dict) -> Optional[dict]:
        """Actualizar proveedor"""
        try:
            if not ObjectId.is_valid(proveedor_id):
                return None
            
            # Filtrar campos None
            update_dict = {k: v for k, v in update_data.items() if v is not None}
            
            if not update_dict:
                return self.get_proveedor_by_id(proveedor_id)
            
            # Si se actualiza el documento, verificar que no exista otro proveedor con el mismo
            if "numero_documento" in update_dict or "tipo_documento" in update_dict:
                proveedor_actual = self.get_proveedor_by_id(proveedor_id)
                tipo_doc = update_dict.get("tipo_documento", proveedor_actual.get("tipo_documento"))
                num_doc = update_dict.get("numero_documento", proveedor_actual.get("numero_documento"))
                
                existing = self.collection.find_one({
                    "tipo_documento": tipo_doc,
                    "numero_documento": num_doc,
                    "_id": {"$ne": ObjectId(proveedor_id)}
                }) 
                
                if existing:
                    raise ValueError(f"Ya existe un proveedor con el {tipo_doc} {num_doc}")
            
            # Actualizar en base de datos
            self.collection.update_one(
                {"_id": ObjectId(proveedor_id)},
                {"$set": update_dict}
            )
            
            return self.get_proveedor_by_id(proveedor_id)
            
        except Exception as e:
            logger.error(f"Error al actualizar proveedor: {str(e)}")
            raise
    
    def delete_proveedor(self, proveedor_id: str) -> bool:
        """Eliminar proveedor"""
        try:
            if not ObjectId.is_valid(proveedor_id):
                return False
            
            result = self.collection.delete_one({"_id": ObjectId(proveedor_id)})
            return result.deleted_count > 0
            
        except Exception as e:
            logger.error(f"Error al eliminar proveedor: {str(e)}")
            return False
    
    def get_all_proveedores_sin_paginacion(self, filter_params: Optional[ProveedorFilter] = None) -> List[dict]:
        """Obtener TODOS los proveedores sin paginación (para exportación)"""
        try:
            query = {}
            
            if filter_params:
                if filter_params.codigo_proveedor:
                    query["codigo_proveedor"] = safe_regex(filter_params.codigo_proveedor)
                
                if filter_params.tipo_documento:
                    query["tipo_documento"] = filter_params.tipo_documento
                
                if filter_params.numero_documento:
                    query["numero_documento"] = safe_regex(filter_params.numero_documento)
                
                if filter_params.razon_social:
                    query["razon_social"] = safe_regex(filter_params.razon_social)
                
                if filter_params.rubro_proveedor:
                    query["rubro_proveedor"] = filter_params.rubro_proveedor
                
                if filter_params.contacto_principal:
                    query["contacto_principal"] = safe_regex(filter_params.contacto_principal)
                
                if filter_params.telefono:
                    query["telefono"] = safe_regex(filter_params.telefono)
                
                if filter_params.estado:
                    query["estado"] = filter_params.estado
                
                if filter_params.servicio:
                    query["servicios"] = safe_regex(filter_params.servicio)
            
            # Limpiar filtros nulos
            query = {k: v for k, v in query.items() if v is not None}
            
            proveedores = list(self.collection.find(query).sort("razon_social", 1))
            
            # Convertir ObjectId a string
            for proveedor in proveedores:
                proveedor["id"] = str(proveedor["_id"])
                del proveedor["_id"]
            
            return proveedores
            
        except Exception as e:
            logger.error(f"Error al obtener proveedores: {str(e)}")
            raise
    
    def export_to_excel(self, filter_params: Optional[ProveedorFilter] = None) -> BytesIO:
        """Exportar proveedores a Excel (formato simplificado)"""
        try:
            proveedores = self.get_all_proveedores_sin_paginacion(filter_params)
            
            if not proveedores:
                # Crear DataFrame vacío con columnas simples
                df = pd.DataFrame(columns=[
                    "Código Proveedor", "Tipo Documento", "Número Documento",
                    "Razón Social", "Rubro Proveedor", "Servicios",
                    "Contacto Principal", "Teléfono", "Email", "Dirección",
                    "Website", "Estado", "Fecha Registro", "Observaciones"
                ])
            else:
                # Preparar datos para Excel (solo campos simples)
                excel_data = []
                for proveedor in proveedores:
                    # Convertir servicios (lista) a string separado por comas
                    servicios_str = ""
                    if proveedor.get("servicios"):
                        servicios_str = ", ".join(proveedor["servicios"])
                    
                    excel_data.append({
                        "Código Proveedor": proveedor.get("codigo_proveedor", ""),
                        "Tipo Documento": proveedor.get("tipo_documento", ""),
                        "Número Documento": proveedor.get("numero_documento", ""),
                        "Razón Social": proveedor.get("razon_social", ""),
                        "Rubro Proveedor": proveedor.get("rubro_proveedor", ""),
                        "Servicios": servicios_str,
                        "Contacto Principal": proveedor.get("contacto_principal", ""),
                        "Teléfono": proveedor.get("telefono", ""),
                        "Email": proveedor.get("email", ""),
                        "Dirección": proveedor.get("direccion", ""),
                        "Website": proveedor.get("website", ""),
                        "Estado": proveedor.get("estado", ""),
                        "Fecha Registro": proveedor.get("fecha_registro", "").strftime("%Y-%m-%d %H:%M:%S") if proveedor.get("fecha_registro") else "",
                        "Observaciones": proveedor.get("observaciones", "")
                    })
                
                df = pd.DataFrame(excel_data)
            
            # Crear Excel en memoria con formato mejorado
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Proveedores')
                
                # Ajustar anchos de columna
                worksheet = writer.sheets['Proveedores']
                for idx, col in enumerate(df.columns, 1):
                    max_length = max(
                        df[col].astype(str).apply(len).max(),
                        len(col)
                    ) + 2
                    worksheet.column_dimensions[chr(64 + idx)].width = min(max_length, 50)
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al exportar a Excel: {str(e)}")
            raise
    
    def import_from_excel(self, file_content: bytes) -> Dict[str, Any]:
        """Importar proveedores desde Excel (formato simplificado)"""
        try:
            import io
            
            df = pd.read_excel(io.BytesIO(file_content))
            
            created = 0
            updated = 0
            errors = []
            skipped = 0
            
            for index, row in df.iterrows():
                try:
                    # Construir datos del proveedor (solo campos simples)
                    proveedor_data = {
                        "tipo_documento": str(row.get("Tipo Documento", "")).strip(),
                        "numero_documento": str(row.get("Número Documento", "")).strip(),
                        "razon_social": str(row.get("Razón Social", "")).strip(),
                        "estado": str(row.get("Estado", "activo")).strip(),
                        "fecha_registro": datetime.now()
                    }
                    
                    # Campos opcionales simples
                    campos_opcionales = {
                        "rubro_proveedor": "Rubro Proveedor",
                        "contacto_principal": "Contacto Principal",
                        "telefono": "Teléfono",
                        "email": "Email",
                        "direccion": "Dirección",
                        "website": "Website",
                        "observaciones": "Observaciones"
                    }
                    
                    for campo_db, campo_excel in campos_opcionales.items():
                        valor = row.get(campo_excel)
                        if pd.notna(valor) and str(valor).strip() and str(valor).strip().lower() != "nan":
                            proveedor_data[campo_db] = str(valor).strip()
                    
                    # Procesar servicios (convertir string separado por comas a lista)
                    servicios_str = row.get("Servicios")
                    if pd.notna(servicios_str) and str(servicios_str).strip() and str(servicios_str).strip().lower() != "nan":
                        servicios_list = [s.strip() for s in str(servicios_str).split(",") if s.strip()]
                        if servicios_list:
                            proveedor_data["servicios"] = servicios_list
                    
                    # Validar campos obligatorios
                    if not proveedor_data.get("tipo_documento"):
                        errors.append(f"Fila {index + 2}: Tipo de documento es requerido")
                        continue
                    
                    if not proveedor_data.get("numero_documento"):
                        errors.append(f"Fila {index + 2}: Número de documento es requerido")
                        continue
                    
                    if not proveedor_data.get("razon_social"):
                        errors.append(f"Fila {index + 2}: Razón social es requerida")
                        continue
                    
                    # Asegurar estado válido
                    if proveedor_data.get("estado") not in ["activo", "inactivo", "suspendido"]:
                        proveedor_data["estado"] = "activo"
                    
                    # Verificar si ya existe por código (si viene en el Excel)
                    codigo_excel = str(row.get("Código Proveedor", "")).strip()
                    if codigo_excel and codigo_excel not in ["", "nan", "None"]:
                        existing = self.collection.find_one({"codigo_proveedor": codigo_excel})
                        if existing:
                            # Actualizar proveedor existente
                            proveedor_data.pop("fecha_registro", None)  # No actualizar fecha de registro
                            self.collection.update_one(
                                {"codigo_proveedor": codigo_excel},
                                {"$set": proveedor_data}
                            )
                            updated += 1
                            continue
                    
                    # Verificar si existe por documento
                    existing_doc = self.collection.find_one({
                        "tipo_documento": proveedor_data["tipo_documento"],
                        "numero_documento": proveedor_data["numero_documento"]
                    })
                    
                    if existing_doc:
                        # Proveedor duplicado, lo saltamos
                        skipped += 1
                        errors.append(f"Fila {index + 2}: Proveedor duplicado - {proveedor_data['tipo_documento']} {proveedor_data['numero_documento']} ya existe")
                        continue
                    
                    # Crear nuevo proveedor (sin código, se generará automáticamente)
                    self.create_proveedor(proveedor_data)
                    created += 1
                        
                except Exception as e:
                    errors.append(f"Fila {index + 2}: {str(e)}")
                    continue
            
            return {
                "total_rows": len(df),
                "created": created,
                "updated": updated,
                "skipped": skipped,
                "errors": errors,
                "has_errors": len(errors) > 0,
                "success_rate": f"{((created + updated) / len(df) * 100):.1f}%" if len(df) > 0 else "0%"
            }
            
        except Exception as e:
            logger.error(f"Error al importar desde Excel: {str(e)}")
            raise

    def generate_excel_template(self) -> BytesIO:
        """Generar plantilla de Excel vacía para importación de proveedores"""
        try:
            # Crear DataFrame con columnas y una fila de ejemplo
            template_data = [{
                "Tipo Documento": "RUC",
                "Número Documento": "20987654321",
                "Razón Social": "TRANSPORTES EJEMPLO SAC",
                "Rubro Proveedor": "transportista",
                "Servicios": "Transporte de carga, Mudanzas, Logística",
                "Contacto Principal": "Carlos Rodríguez",
                "Teléfono": "987654321",
                "Email": "contacto@ejemplo.com",
                "Dirección": "Av. Los Transportistas 456, Lima",
                "Website": "www.ejemplo.com",
                "Estado": "activo",
                "Observaciones": "Proveedor ejemplo - puede eliminar esta fila"
            }]
            
            df = pd.DataFrame(template_data)
            
            # Crear Excel en memoria con formato
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Proveedores')
                
                workbook = writer.book
                worksheet = writer.sheets['Proveedores']
                
                # Estilo para encabezados
                from openpyxl.styles import Font, PatternFill, Alignment
                
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Ajustar anchos de columna
                column_widths = {
                    'A': 18,  # Tipo Documento
                    'B': 18,  # Número Documento
                    'C': 35,  # Razón Social
                    'D': 20,  # Rubro Proveedor
                    'E': 50,  # Servicios
                    'F': 25,  # Contacto Principal
                    'G': 15,  # Teléfono
                    'H': 30,  # Email
                    'I': 40,  # Dirección
                    'J': 25,  # Website
                    'K': 12,  # Estado
                    'L': 40   # Observaciones
                }
                
                for col, width in column_widths.items():
                    worksheet.column_dimensions[col].width = width
                
                # Agregar instrucciones en una hoja separada
                instructions_data = {
                    "Campo": [
                        "Tipo Documento",
                        "Número Documento",
                        "Razón Social",
                        "Estado",
                        "Rubro Proveedor",
                        "Servicios",
                        "Contacto Principal",
                        "Teléfono",
                        "Email",
                        "Dirección",
                        "Website",
                        "Observaciones"
                    ],
                    "Obligatorio": [
                        "SÍ", "SÍ", "SÍ", "NO",
                        "NO", "NO", "NO", "NO",
                        "NO", "NO", "NO", "NO"
                    ],
                    "Descripción": [
                        "Tipo de documento: RUC, DNI, CE",
                        "Número del documento de identidad",
                        "Nombre o razón social del proveedor",
                        "Estado del proveedor: activo, inactivo, suspendido (por defecto: activo)",
                        "Rubro: transportista, logistica, seguridad, mantenimiento, tecnologia, seguros, servicios, otros",
                        "Lista de servicios separados por comas",
                        "Nombre del contacto principal",
                        "Número de teléfono",
                        "Correo electrónico",
                        "Dirección completa",
                        "Sitio web",
                        "Observaciones o notas adicionales"
                    ],
                    "Ejemplo": [
                        "RUC",
                        "20987654321",
                        "TRANSPORTES EJEMPLO SAC",
                        "activo",
                        "transportista",
                        "Transporte de carga, Mudanzas, Logística",
                        "Carlos Rodríguez",
                        "987654321",
                        "contacto@ejemplo.com",
                        "Av. Los Transportistas 456, Lima",
                        "www.ejemplo.com",
                        "Proveedor confiable"
                    ]
                }
                
                df_instructions = pd.DataFrame(instructions_data)
                df_instructions.to_excel(writer, sheet_name='Instrucciones', index=False)
                
                # Formatear hoja de instrucciones
                ws_instructions = writer.sheets['Instrucciones']
                
                for cell in ws_instructions[1]:
                    cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                ws_instructions.column_dimensions['A'].width = 30
                ws_instructions.column_dimensions['B'].width = 15
                ws_instructions.column_dimensions['C'].width = 60
                ws_instructions.column_dimensions['D'].width = 40
                
                # Ajustar altura de filas en instrucciones
                for row in ws_instructions.iter_rows(min_row=2, max_row=ws_instructions.max_row):
                    ws_instructions.row_dimensions[row[0].row].height = 30
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="center")
            
            output.seek(0)
            return output
            
        except Exception as e:
            logger.error(f"Error al generar plantilla Excel: {str(e)}")
            raise

    def get_stats(self) -> Dict[str, Any]:
        """Obtener estadísticas de proveedores"""
        try:
            total = self.collection.count_documents({})
            activos = self.collection.count_documents({"estado": "activo"})
            inactivos = self.collection.count_documents({"estado": "inactivo"})
            suspendidos = self.collection.count_documents({"estado": "suspendido"})
            
            # Agrupar por rubro de proveedor
            pipeline_rubro = [
                {"$group": {"_id": "$rubro_proveedor", "count": {"$sum": 1}}}
            ]
            
            rubros = {}
            for result in self.collection.aggregate(pipeline_rubro):
                rubros[result["_id"] if result["_id"] else "Sin especificar"] = result["count"]
            
            # Agrupar por tipo de documento
            pipeline_doc = [
                {"$group": {"_id": "$tipo_documento", "count": {"$sum": 1}}}
            ]
            
            documentos = {}
            for result in self.collection.aggregate(pipeline_doc):
                documentos[result["_id"]] = result["count"]
            
            # Servicios más comunes (desenrollar array)
            pipeline_servicios = [
                {"$unwind": "$servicios"},
                {"$group": {"_id": "$servicios", "count": {"$sum": 1}}},
                {"$sort": {"count": -1}},
                {"$limit": 10}
            ]
            
            servicios = {}
            for result in self.collection.aggregate(pipeline_servicios):
                if result["_id"]:
                    servicios[result["_id"]] = result["count"]
            
            return {
                "total": total,
                "activos": activos,
                "inactivos": inactivos,
                "suspendidos": suspendidos,
                "por_rubro": rubros,
                "por_tipo_documento": documentos,
                "servicios_mas_comunes": servicios
            }
            
        except Exception as e:
            logger.error(f"Error al obtener estadísticas: {str(e)}")
            return {}


def safe_regex(value: str):
    """
    Función auxiliar para crear expresiones regulares seguras en MongoDB
    Escapa caracteres especiales y retorna None si el valor es vacío
    """
    if not value:
        return None
    value = value.strip()
    if not value:
        return None
    return {"$regex": re.escape(value), "$options": "i"}